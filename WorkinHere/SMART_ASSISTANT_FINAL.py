"""
SMART Assistant v0.2.42
User Management & Automation Tool
"""

import tkinter as tk
from tkinter import messagebox, ttk, scrolledtext, filedialog
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
import time
import threading
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import json
import logging
from datetime import datetime as dt_now

# Auto-updater (only if running as EXE)
AUTO_UPDATE_ENABLED = False
try:
    from auto_updater import AutoUpdater
    AUTO_UPDATE_ENABLED = True
    print("[STARTUP] Auto-updater module loaded successfully")
except ImportError as e:
    print(f"[STARTUP] Auto-updater not available: {str(e)}")
    print("[STARTUP] This is normal when running from source (.py file)")
except Exception as e:
    print(f"[STARTUP] ERROR loading auto-updater: {str(e)}")
    import traceback
    traceback.print_exc()

# ================= APP INFO =================
APP_VERSION = "0.2.42"
APP_NAME = "SMART Assistant"

# ================= LOGGING SETUP =================
import logging
import sys
import os

# Configure logging to both file and console
log_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
log_file = os.path.join(log_dir, "smart_assistant.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, mode='a'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)
logger.info(f"="*60)
logger.info(f"SMART Assistant v{APP_VERSION} Starting...")
logger.info(f"Log file: {log_file}")
logger.info(f"="*60)

# ================= GITHUB CONFIG =================
# GitHub configuration for auto-updates
GITHUB_USER = "ohhhthehorror"
GITHUB_REPO = "SMART-Assistant-Publish"

# ================= SCAN CONFIGURATION =================
# Environment URLs
PRODUCTION_URL = "https://fe.blockbyblock.com/auth/login"
PREPROD_URL = "https://preprodfe.blockbyblock.com/auth/login"

# Default to production (will be set by user preference)
LOGIN_URL = PRODUCTION_URL

INACTIVE_DAYS = 120

# Helper function to get environment-aware URL
def get_url(path, app=None):
    """
    Get URL for given path based on environment setting
    Args:
        path: URL path (e.g., '/main/users', '/auth/login')
        app: App instance to check environment setting
    Returns:
        Full URL with correct domain based on environment
    """
    if app and hasattr(app, 'environment') and app.environment == "PreProd":
        base = "https://preprodfe.blockbyblock.com"
    else:
        base = "https://fe.blockbyblock.com"
    
    return f"{base}{path}"

TABLE_SELECTOR = "table"
ROWS_SELECTOR  = "tbody tr"

COL_NAME_LINK    = 1
COL_JOB_TITLE    = 2
COL_USERNAME     = 3
COL_LAST_LOGIN   = 6

DATE_FORMAT = "%m/%d/%Y %I:%M %p"

# Pagination settings
PAGINATION_TAB_COUNT = 2
PAGE_SELECT_ID = "demo-select-small"
FIRST_PAGE_VALUE = "01"

# Checkbox IDs for deactivation
CHECKBOX_IDS = ["activated", "shiftEntry", "activeStatus"]

# Save button selector
SAVE_BUTTON_SELECTOR = (
    By.XPATH,
    "//button[contains(text(), 'SAVE')]"
)

# ================= CORE DEACTIVATION FUNCTION =================

def deactivate_user_by_ticket(driver, username, ticket_number, automation_speed=1.0, app=None):
    """
    Deactivate a single user with ticket number
    Returns: True if successful, False if failed
    automation_speed: multiplier for sleep times (default 1.0)
    app: App instance for environment-aware URLs
    Base timings: Tabs 0.5s, Text 0.1s, Clicks 1s, Pages 3s
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    import time
    
    wait = WebDriverWait(driver, 10)
    actions = ActionChains(driver)
    
    try:
        print(f"\n→ Deactivating user: {username} with Ticket: {ticket_number}")
        
        # Navigate to users page
        print("→ Loading users page...")
        driver.get(get_url("/main/users", app))
        time.sleep(3.0 / automation_speed)  # Page load
        driver.find_element(By.TAG_NAME, "body").click()
        time.sleep(1.0 / automation_speed)  # Click
        
        # Search for user
        print(f"→ Searching for: {username}")
        driver.execute_script("""
            var searchInput = document.querySelector('.table-search-input');
            if (searchInput) {
                searchInput.focus();
                searchInput.click();
            }
        """)
        time.sleep(0.5 / automation_speed)  # Tab/focus
        
        # Type username and search
        actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(username).perform()
        time.sleep(0.1 / automation_speed)  # Text entry
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(3.0 / automation_speed)  # Page load/search
        
        # Find and click user link
        user_link = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "span.text-link")
        ))
        print("→ Opening user profile...")
        driver.execute_script("arguments[0].click();", user_link)
        time.sleep(3.0 / automation_speed)  # Page load
        
        # Execute deactivation sequence
        print(f"→ Executing deactivation with Ticket {ticket_number}...")
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.ARROW_DOWN).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.SPACE).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(f"Ticket {ticket_number} Deactivated").perform()
        time.sleep(0.1 / automation_speed)  # Text entry
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys("12@Deactivate").perform()
        time.sleep(0.1 / automation_speed)  # Text entry
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.SPACE).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.SPACE).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.SPACE).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(0.5 / automation_speed)  # Keypress
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5 / automation_speed)  # Tab
        actions.send_keys("12@Deactivate").perform()
        time.sleep(0.1 / automation_speed)  # Text entry
        
        # Scroll and save
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.5 / automation_speed)  # Scroll wait
        
        save_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(text(), 'SAVE')]")
        ))
        driver.execute_script("arguments[0].click();", save_button)
        time.sleep(3.0 / automation_speed)  # Page load/save
        
        print(f"✓ User '{username}' deactivated successfully!")
        return True
        
    except Exception as e:
        print(f"✗ Deactivation failed: {str(e)}")
        return False

# ================= LOGGING SYSTEM =================

def setup_logging():
    """Setup file logging and crash reporting - NO CMD WINDOW"""
    try:
        # Get script directory
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
        else:
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        
        # Create logs directory
        log_dir = os.path.join(script_dir, "logs")
        os.makedirs(log_dir, exist_ok=True)
        
        # Create log filename with timestamp
        timestamp = dt_now.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_file = os.path.join(log_dir, f"smart_assistant_{timestamp}.log")
        
        # Keep only last 10 log files
        try:
            log_files = sorted([f for f in os.listdir(log_dir) if f.endswith('.log')])
            if len(log_files) > 10:
                for old_file in log_files[:-10]:
                    os.remove(os.path.join(log_dir, old_file))
        except:
            pass
        
        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s] %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)  # Still show in console if opened manually
            ]
        )
        
        # Redirect print statements to logging
        class LogPrinter:
            def write(self, message):
                if message.strip():
                    logging.info(message.strip())
            def flush(self):
                pass
        
        # Only redirect stdout/stderr if not in console mode
        # This allows manual debugging while keeping normal runs clean
        if not sys.stdout.isatty():
            sys.stdout = LogPrinter()
            sys.stderr = LogPrinter()
        
        logging.info("="*60)
        logging.info("SMART Assistant v2.0.5 Starting")
        logging.info(f"Log file: {log_file}")
        logging.info("="*60)
        
        return log_file
    except Exception as e:
        # If logging setup fails, continue anyway
        print(f"Warning: Could not setup logging: {e}")
        return None

# Setup logging immediately
CURRENT_LOG_FILE = setup_logging()

# ================= CRASH REPORTER =================

def save_crash_report(error_info):
    """Save crash report as JSON for easy debugging"""
    try:
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
        else:
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        
        crash_file = os.path.join(script_dir, "last_crash.json")
        
        crash_data = {
            "timestamp": dt_now.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error_type": error_info.get("type", "Unknown"),
            "error_message": error_info.get("message", ""),
            "traceback": error_info.get("traceback", ""),
            "log_file": CURRENT_LOG_FILE
        }
        
        with open(crash_file, 'w') as f:
            json.dump(crash_data, f, indent=4)
        
        logging.error("="*60)
        logging.error("CRASH REPORT SAVED")
        logging.error(f"Location: {crash_file}")
        logging.error("="*60)
        
    except:
        pass

# ================= TIMING MANAGER =================

class TimingManager:
    """Centralized timing manager for all progress bars"""
    
    def __init__(self):
        self.timing_file = "smart_assistant_timings.json"
        self.timings = self.load_timings()
    
    def load_timings(self):
        """Load timing averages from file"""
        defaults = {
            "splash_screen": 5.0,
            "browser_init": 15.0,
            "username_scan": 2.5,  # per page
            "history_scan": 2.2,   # per user
            "deactivation": 8.0,   # per user
            "single_deactivation": 8.0
        }
        
        try:
            if os.path.exists(self.timing_file):
                with open(self.timing_file, 'r') as f:
                    loaded = json.load(f)
                    defaults.update(loaded)
        except:
            pass
        
        return defaults
    
    def save_timings(self):
        """Save timing averages to file"""
        try:
            with open(self.timing_file, 'w') as f:
                json.dump(self.timings, f, indent=2)
        except:
            pass
    
    def update_timing(self, task_name, duration, count=1):
        """Update average timing for a task"""
        if task_name not in self.timings:
            self.timings[task_name] = duration
        else:
            # Running average
            current = self.timings.get(task_name + "_count", 0)
            total = self.timings[task_name] * current + duration * count
            new_count = current + count
            self.timings[task_name] = total / new_count
            self.timings[task_name + "_count"] = new_count
        
        self.save_timings()
    
    def get_estimate(self, task_name, count=1):
        """Get estimated time for a task"""
        base_time = self.timings.get(task_name, 1.0)
        return base_time * count

# Global timing manager instance
timing_manager = TimingManager()



# ================= CONFIGURATION =================

LOGIN_URL = "https://fe.blockbyblock.com/auth/login"

INACTIVE_DAYS = 120

TABLE_SELECTOR = "table"
ROWS_SELECTOR  = "tbody tr"

COL_NAME_LINK    = 1
COL_JOB_TITLE    = 2
COL_USERNAME     = 3
COL_LAST_LOGIN   = 6

DATE_FORMAT = "%m/%d/%Y %I:%M %p"

# Navigation constants
PAGINATION_TAB_COUNT = 2
PAGE_SELECT_ID = "demo-select-small"
FIRST_PAGE_VALUE = "01"

CHECKBOX_IDS = ["activated", "shiftEntry", "activeStatus"]

SAVE_BUTTON_SELECTOR = (
    "button[type='submit'], button.save, #save, .btn-primary, "
    "[type='submit'], button:contains('Save')"
)

# ======================================================================

class SplashScreen:
    def __init__(self):
        import json
        import os
        import sys
        
        # Get script directory properly
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
        else:
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        
        pref_file = os.path.join(script_dir, "smart_assistant_settings.json")
        self.is_dark_mode = False  # DEFAULT LIGHT!
        
        print("\n" + "="*60)
        print("SPLASH SCREEN STARTING")
        print(f"Theme file: {pref_file}")
        
        try:
            if os.path.exists(pref_file):
                with open(pref_file, 'r') as f:
                    data = json.load(f)
                    self.is_dark_mode = data.get("is_dark_mode", False)
                    print(f"Theme: {'DARK' if self.is_dark_mode else 'LIGHT'}")
            else:
                print("No theme file - using LIGHT mode")
        except Exception as e:
            print(f"Error loading theme: {e}")
        
        print("="*60 + "\n")
        
        self.root = tk.Tk()
        self.root.overrideredirect(True)
        
        # Update to get screen info before positioning
        self.root.update_idletasks()
        
        # Original size
        width = 750
        height = 200
        
        # Center on monitor where mouse is
        x, y = self.get_monitor_center(width, height)
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.attributes('-topmost', True)
        
        # Modern theme colors
        if self.is_dark_mode:
            bg = "#0f0f0f"           # Almost black (modern)
            border = "#0052CC"       # Bright cyan
            text = "#ffffff"         # White
            progress_bg = "#252525"  # Elevated surface
        else:
            bg = "#f5f3f0"           # Warm off-white/cream
            border = "#0066ff"       # Vibrant blue
            text = "#1a1a1a"         # Near black
            progress_bg = "#fafaf8"  # Light cream elevated surface
        
        # Border
        border_frame = tk.Frame(self.root, bg=border, bd=0)
        border_frame.pack(fill="both", expand=True)
        
        inner = tk.Frame(border_frame, bg=bg, bd=0)
        inner.pack(fill="both", expand=True, padx=3, pady=3)
        
        # Just header
        tk.Label(inner, text="SMART Assistant", font=("Segoe UI", 32, "bold"), bg=bg, fg=border).pack(pady=(30, 10))
        tk.Label(inner, text="Version 2.0.5", font=("Segoe UI", 12, "bold"), bg=bg, fg=border).pack(pady=(0, 20))
        
        # Progress
        prog_container = tk.Frame(inner, bg=bg)
        prog_container.pack(fill="x", padx=100, pady=(10, 20))
        
        prog_bg_frame = tk.Frame(prog_container, bg=progress_bg, height=12)
        prog_bg_frame.pack(fill="x")
        
        self.progress_bar = tk.Frame(prog_bg_frame, bg=border, height=12)
        self.progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        self.loading_label = tk.Label(inner, text="Loading...", font=("Segoe UI", 10), bg=bg, fg=text)
        self.loading_label.pack()
        
        self.root.update()
    
    def get_monitor_center(self, width, height):
        """Get center coordinates for the monitor where mouse is located"""
        try:
            # Update to get accurate mouse position
            self.root.update_idletasks()
            
            # Get mouse position
            mouse_x = self.root.winfo_pointerx()
            mouse_y = self.root.winfo_pointery()
            
            print(f"Mouse position: {mouse_x}, {mouse_y}")
            
            # Get screen dimensions (total virtual screen)
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            print(f"Screen dimensions: {screen_width} x {screen_height}")
            
            # Try to detect monitor using tkinter's multi-monitor support
            # Most common setup: monitors side-by-side at 1920px each
            # If screen_width > screen_height, assume horizontal arrangement
            if screen_width > 2000:  # Multiple monitors likely
                # Estimate number of monitors (assume 1920px each)
                num_monitors = round(screen_width / 1920)
                monitor_width = screen_width / num_monitors
                
                print(f"Detected {num_monitors} monitors, width={monitor_width}")
                
                # Which monitor is mouse on?
                monitor_index = int(mouse_x / monitor_width)
                
                print(f"Mouse on monitor {monitor_index}")
                
                # Get bounds of that monitor
                monitor_left = monitor_index * monitor_width
                monitor_center_x = monitor_left + (monitor_width / 2)
                monitor_center_y = screen_height / 2
            else:
                # Single monitor or can't detect - use simple center
                monitor_center_x = screen_width / 2
                monitor_center_y = screen_height / 2
                print("Single monitor detected")
            
            # Center the window on that monitor
            x = int(monitor_center_x - (width / 2))
            y = int(monitor_center_y - (height / 2))
            
            print(f"Centering at: {x}, {y}")
            
            return x, y
        except Exception as e:
            print(f"Monitor detection error: {e}")
            # Fallback to simple center
            return (self.root.winfo_screenwidth() // 2) - (width // 2), (self.root.winfo_screenheight() // 2) - (height // 2)
    
    def load_theme_preference(self):
        """Load theme preference - kept for compatibility"""
        return self.is_dark_mode
    
    def animate_smooth(self, duration=1.0):
        """Smooth animation - 1 second"""
        import time
        steps = 100
        
        for i in range(steps + 1):
            progress = i / 100.0
            self.progress_bar.place(relwidth=progress)
            self.loading_label.config(text=f"Loading... {int(progress * 100)}%")
            self.root.update()
            time.sleep(duration / steps)
    
    def close(self):
        self.root.destroy()

# ======================================================================
# ======================================================================


def check_internet_connection():
    """Check if there's an active internet connection"""
    import socket
    try:
        # Try to connect to Google DNS
        socket.create_connection(("8.8.8.8", 53), timeout=3)
        return True
    except OSError:
        return False


def set_window_icon(root_window):
    """Helper function to set window icon - call AFTER window is fully configured"""
    logging.info("[ICON] Setting window icon...")
    try:
        import os
        import sys
        
        # Set Windows taskbar icon (fixes taskbar not showing icon)
        if os.name == 'nt':
            import ctypes
            myappid = 'SMS.SMARTAssistant.v2.0'  # arbitrary string
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            logging.info("[ICON] Set Windows AppUserModelID")
        
        from PIL import Image, ImageTk
        
        icon_paths = [
            r"C:\Users\F_Johnson\OneDrive - SMS Holdings Corporation\Desktop\Programstest\smartlogosmall.png",
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "smartlogosmall.png"),
            os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__), "smartlogosmall.png")
        ]
        
        logging.info(f"[ICON] Trying {len(icon_paths)} icon paths...")
        for i, icon_path in enumerate(icon_paths):
            logging.info(f"[ICON] Path {i+1}: {icon_path}")
            if os.path.exists(icon_path):
                logging.info(f"[ICON] ✅ Found icon at path {i+1}")
                try:
                    # For Windows, try .ico first for taskbar
                    if os.name == 'nt' and icon_path.endswith('.png'):
                        ico_path = icon_path.replace('.png', '.ico')
                        if os.path.exists(ico_path):
                            root_window.iconbitmap(ico_path)
                            logging.info(f"[ICON] ✅ Used .ico file for Windows taskbar")
                    
                    # Then set PNG for window icon
                    img = Image.open(icon_path)
                    logging.info(f"[ICON] ✅ Loaded image: {img.size} {img.format}")
                    
                    # Resize icon to proper size for window icons (32x32)
                    if img.size[0] > 64 or img.size[1] > 64:
                        img = img.resize((32, 32), Image.Resampling.LANCZOS)
                        logging.info(f"[ICON] ✅ Resized to 32x32 for better quality")
                    
                    # Create PhotoImage with the correct parent window
                    icon_img = ImageTk.PhotoImage(img, master=root_window)
                    logging.info(f"[ICON] ✅ Created PhotoImage")
                    
                    root_window.iconphoto(True, icon_img)
                    logging.info(f"[ICON] ✅ Called iconphoto()")
                    root_window._icon_img = icon_img  # Keep reference
                    logging.info(f"[ICON] ✅ Stored reference")
                    return True
                except ImportError as e:
                    logging.info(f"[ICON] ❌ PIL Import error: {e}")
                    if icon_path.endswith('.png') and os.name == 'nt':
                        ico_path = icon_path.replace('.png', '.ico')
                        if os.path.exists(ico_path):
                            root_window.iconbitmap(ico_path)
                            logging.info(f"[ICON] ✅ Used ICO fallback")
                            return True
                except Exception as e:
                    logging.info(f"[ICON] ❌ Error: {e}")
            else:
                logging.info(f"[ICON] ❌ Not found at path {i+1}")
    except Exception as e:
        logging.info(f"[ICON] ❌ Fatal error: {e}")
    logging.info("[ICON] ❌ Failed to set icon")
    return False




# ======================================================================
# SINGLE-WINDOW APPLICATION
# ======================================================================

class SMARTAssistantApp:
    """Main application container - ONE window"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"{APP_NAME} v{APP_VERSION}")
        
        # Position at mouse cursor
        try:
            mouse_x = self.root.winfo_pointerx()
            mouse_y = self.root.winfo_pointery()
            self.root.geometry(f"+{mouse_x}+{mouse_y}")
            self.root.update()
            self.root.state('zoomed')
        except:
            self.root.state('zoomed')
        
        set_window_icon(self.root)
        
        # Load theme
        self.is_dark_mode = self.load_theme_preference()
        self.apply_theme_colors()
        self.root.configure(bg=self.bg_dark)
        
        # Enable dark title bar if in dark mode
        self.enable_dark_title_bar(self.root)
        
        # Container for swapping screens
        self.container = tk.Frame(self.root, bg=self.bg_dark)
        self.container.pack(fill="both", expand=True)
        
        self.current_screen = None
        self.username = None
        self.password = None
        
        # Load headless mode from settings (must be done AFTER load_theme_preference created the method)
        # Default to True if not in settings
        self.headless_mode = True  # Default
        self.automation_speed = 1.0  # Default speed multiplier (1x = normal speed)
        self.environment = "Production"  # Default environment
        self.custom_filters = []  # User-added name filters
        
        logger.info("Loading settings...")
        self.load_all_settings()  # This will override with saved values if they exist
        logger.info(f"Settings loaded - Headless: {self.headless_mode}, Speed: {self.automation_speed}, Environment: {self.environment}")
        
        self.driver = None  # Chrome driver instance
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Check for updates BEFORE showing login
        if AUTO_UPDATE_ENABLED:
            logger.info("Auto-updater enabled")
            logger.info(f"Current version: {APP_VERSION}")
            logger.info(f"GitHub: {GITHUB_USER}/{GITHUB_REPO}")
            print(f"[DEBUG] Auto-updater enabled")
            print(f"[DEBUG] Current version: {APP_VERSION}")
            print(f"[DEBUG] GitHub: {GITHUB_USER}/{GITHUB_REPO}")
            
            self.updater = AutoUpdater(APP_VERSION, GITHUB_USER, GITHUB_REPO)
            # Pass theme info to updater
            self.updater.is_dark_mode = self.is_dark_mode
            self.updater.bg_card = self.bg_card
            self.updater.text_light = self.text_light
            self.updater.accent = self.accent
            
            # Check for updates in background, then show login
            def check_then_login():
                logger.info("Starting update check...")
                print("[DEBUG] Starting update check...")
                self.updater.check_and_prompt_update(self.root)
                # Show login after a short delay (whether update found or not)
                print("[DEBUG] Update check complete, showing login...")
                self.root.after(1000, self.show_login_screen)
            
            self.root.after(500, check_then_login)
        else:
            print("[DEBUG] Auto-updater NOT enabled")
            # No updater, just show login
            self.show_login_screen()
    
    def load_theme_preference(self):
        """Load theme preference from settings file"""
        import json, os, sys
        try:
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
            else:
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            pref_file = os.path.join(script_dir, "smart_assistant_settings.json")
            if os.path.exists(pref_file):
                with open(pref_file, 'r') as f:
                    settings = json.load(f)
                    return settings.get("is_dark_mode", False)
        except:
            pass
        return False
    
    def load_all_settings(self):
        """Load all settings including headless mode and automation speed"""
        import json, os, sys
        try:
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
            else:
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            pref_file = os.path.join(script_dir, "smart_assistant_settings.json")
            if os.path.exists(pref_file):
                with open(pref_file, 'r') as f:
                    settings = json.load(f)
                    # Load headless mode
                    self.headless_mode = settings.get("headless_mode", True)
                    # Load automation speed
                    self.automation_speed = settings.get("automation_speed", 1.0)
                    # Load environment
                    self.environment = settings.get("environment", "Production")
                    # Load scan timing data
                    self.last_scan_time = settings.get("last_scan_time", None)
                    self.last_scan_pages = settings.get("last_scan_pages", None)
                    # Load custom filters
                    self.custom_filters = settings.get("custom_filters", [])
        except:
            pass
    
    def save_settings(self):
        """Save all settings to file"""
        import json, os, sys
        try:
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
            else:
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            pref_file = os.path.join(script_dir, "smart_assistant_settings.json")
            
            settings = {
                "is_dark_mode": self.is_dark_mode,
                "headless_mode": self.headless_mode,
                "automation_speed": self.automation_speed,
                "environment": self.environment,
                "last_scan_time": getattr(self, 'last_scan_time', None),
                "last_scan_pages": getattr(self, 'last_scan_pages', None),
                "custom_filters": self.custom_filters
            }
            
            with open(pref_file, 'w') as f:
                json.dump(settings, f, indent=2)
            
            logger.info(f"Settings saved to: {pref_file}")
            logger.info(f"  Dark Mode: {self.is_dark_mode}")
            logger.info(f"  Headless: {self.headless_mode}")
            logger.info(f"  Speed: {self.automation_speed}")
            logger.info(f"  Environment: {self.environment}")
        except Exception as e:
            logger.error(f"Error saving settings: {e}")
            print(f"Error saving settings: {e}")
    
    def exclude_user_from_scans(self, name, parent_dialog=None):
        """Add user name to custom filters"""
        if self.show_dark_confirm("Exclude User", 
                                  f"Exclude '{name}' from future scans?\n\nThis name will be added to your custom filter list."):
            if name.lower() not in [f.lower() for f in self.custom_filters]:
                self.custom_filters.append(name)
                self.save_settings()
                self.show_dark_messagebox("Filter Added", 
                                        f"'{name}' has been added to your custom filters.\n\nThis user will be skipped in future scans.",
                                        "info")
                logger.info(f"Added custom filter: {name}")
            else:
                self.show_dark_messagebox("Already Filtered", 
                                        f"'{name}' is already in your custom filters.",
                                        "info")
    
    def show_filter_management_dialog(self):
        """Show dialog to manage scan filters before starting"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Scan Filter Management")
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        dialog.geometry("600x600")  # Increased height for buttons
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"600x600+{x}+{y}")
        
        # Header
        tk.Label(dialog, text="Scan Filters", 
                font=("Segoe UI", 16, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.accent).pack(pady=15, padx=20)
        
        tk.Label(dialog, text="Filters scan Name, Username, and Job Title fields", 
                font=("Segoe UI", 10, "italic"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_dim).pack(pady=(0, 10), padx=20)
        
        # Content frame
        content = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        content.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Built-in filters (read-only)
        tk.Label(content, text="Built-in Filters (cannot be modified):", 
                font=("Segoe UI", 10, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_light, anchor="w").pack(fill="x", pady=(0, 5))
        
        builtin_frame = tk.Frame(content, bg=self.bg_dark if self.is_dark_mode else "#f0f0f0", 
                                relief="solid", borderwidth=1)
        builtin_frame.pack(fill="x", pady=(0, 15))
        
        builtin_filters = ["deactivat", "deact", "terminat", "termin", "end access", "client (job title)"]
        for f in builtin_filters:
            tk.Label(builtin_frame, text=f"  • {f}", 
                    font=("Segoe UI", 9),
                    bg=self.bg_dark if self.is_dark_mode else "#f0f0f0",
                    fg=self.text_dim, anchor="w").pack(fill="x")
        
        # Custom filters (editable)
        tk.Label(content, text="Your Custom Filters:", 
                font=("Segoe UI", 10, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_light, anchor="w").pack(fill="x", pady=(0, 5))
        
        # Custom filter list
        filter_frame = tk.Frame(content, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        filter_frame.pack(fill="both", expand=True)
        
        # Scrollable listbox
        listbox_frame = tk.Frame(filter_frame, bg=self.bg_dark if self.is_dark_mode else "#f0f0f0",
                                relief="solid", borderwidth=1)
        listbox_frame.pack(fill="both", expand=True, side="left")
        
        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side="right", fill="y")
        
        filter_listbox = tk.Listbox(listbox_frame,
                                    yscrollcommand=scrollbar.set,
                                    bg=self.bg_card if self.is_dark_mode else "#ffffff",
                                    fg=self.text_light,
                                    selectbackground=self.accent,
                                    selectforeground="white",
                                    font=("Segoe UI", 9),
                                    borderwidth=0)
        filter_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=filter_listbox.yview)
        
        # Populate listbox
        def refresh_listbox():
            filter_listbox.delete(0, tk.END)
            for f in self.custom_filters:
                filter_listbox.insert(tk.END, f)
        
        refresh_listbox()
        
        # Buttons for managing filters
        button_panel = tk.Frame(filter_frame, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        button_panel.pack(side="right", fill="y", padx=(10, 0))
        
        def add_filter():
            # Create custom themed input dialog
            input_dialog = tk.Toplevel(dialog)
            input_dialog.title("Add Filter")
            input_dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
            input_dialog.transient(dialog)
            input_dialog.grab_set()
            
            if self.is_dark_mode:
                self.enable_dark_title_bar(input_dialog)
            
            input_dialog.geometry("450x220")
            input_dialog.update_idletasks()
            dx = dialog.winfo_x() + (dialog.winfo_width() // 2) - (input_dialog.winfo_width() // 2)
            dy = dialog.winfo_y() + (dialog.winfo_height() // 2) - (input_dialog.winfo_height() // 2)
            input_dialog.geometry(f"450x220+{dx}+{dy}")
            
            tk.Label(input_dialog, text="Enter value to filter", 
                    font=("Segoe UI", 12, "bold"),
                    bg=self.bg_card if self.is_dark_mode else "#ffffff",
                    fg=self.text_light).pack(pady=(20, 5))
            
            tk.Label(input_dialog, text="Name: Lastname, Firstname | Username: jdoe | Job Title: Manager", 
                    font=("Segoe UI", 9, "italic"),
                    bg=self.bg_card if self.is_dark_mode else "#ffffff",
                    fg=self.text_dim).pack(pady=(0, 5))
            
            tk.Label(input_dialog, text="Exact match (case-insensitive)", 
                    font=("Segoe UI", 9, "italic"),
                    bg=self.bg_card if self.is_dark_mode else "#ffffff",
                    fg=self.text_dim).pack(pady=(0, 10))
            
            entry = tk.Entry(input_dialog, font=("Segoe UI", 11),
                           bg=self.bg_dark if self.is_dark_mode else "#f0f0f0",
                           fg=self.text_light,
                           insertbackground=self.text_light,
                           width=35)
            entry.pack(pady=(0, 20))
            entry.focus()
            
            result = {"value": None}
            
            def ok():
                result["value"] = entry.get().strip()
                input_dialog.destroy()
            
            def cancel():
                input_dialog.destroy()
            
            entry.bind("<Return>", lambda e: ok())
            entry.bind("<Escape>", lambda e: cancel())
            
            btn_frame = tk.Frame(input_dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
            btn_frame.pack(pady=10)
            
            tk.Button(btn_frame, text="Cancel", command=cancel,
                     bg="#999999", fg="white", font=("Segoe UI", 10, "bold"),
                     relief="flat", cursor="hand2", padx=20, pady=8).pack(side="left", padx=5)
            
            tk.Button(btn_frame, text="Add", command=ok,
                     bg=self.accent, fg="white", font=("Segoe UI", 10, "bold"),
                     relief="flat", cursor="hand2", padx=20, pady=8).pack(side="left", padx=5)
            
            input_dialog.wait_window()
            new_filter = result["value"]
            
            if new_filter:
                # Check exact match (case-insensitive)
                if new_filter.lower() not in [f.lower() for f in self.custom_filters]:
                    self.custom_filters.append(new_filter)
                    refresh_listbox()
                    self.save_settings()
                    logger.info(f"Added custom filter (exact): {new_filter}")
                else:
                    self.show_dark_messagebox("Already Exists", 
                                            f"'{new_filter}' is already in your filters.",
                                            "info")
        
        def remove_filter():
            selection = filter_listbox.curselection()
            if selection:
                idx = selection[0]
                removed = self.custom_filters.pop(idx)
                refresh_listbox()
                self.save_settings()
                logger.info(f"Removed custom filter: {removed}")
        
        tk.Button(button_panel, text="➕ Add", command=add_filter,
                 bg=self.accent, fg="white", font=("Segoe UI", 9, "bold"),
                 relief="flat", cursor="hand2", padx=15, pady=5).pack(fill="x", pady=(0, 5))
        
        tk.Button(button_panel, text="➖ Remove", command=remove_filter,
                 bg="#ff4444", fg="white", font=("Segoe UI", 9, "bold"),
                 relief="flat", cursor="hand2", padx=15, pady=5).pack(fill="x")
        
        # Bottom buttons
        bottom_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        bottom_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        result = {"start": False}
        
        def start_scan():
            result["start"] = True
            dialog.destroy()
        
        tk.Button(bottom_frame, text="← Back", command=dialog.destroy,
                 bg="#999999", fg="white", font=("Segoe UI", 11, "bold"),
                 relief="flat", cursor="hand2", padx=20, pady=10).pack(side="left")
        
        tk.Button(bottom_frame, text="Start Scan →", command=start_scan,
                 bg="#4CAF50", fg="white", font=("Segoe UI", 11, "bold"),
                 relief="flat", cursor="hand2", padx=20, pady=10).pack(side="right")
        
        dialog.wait_window()
        return result["start"]
    
    def update_login_url(self):
        """Update the LOGIN_URL global variable based on environment setting"""
        global LOGIN_URL
        if self.environment == "Production":
            LOGIN_URL = PRODUCTION_URL
        else:  # PreProd
            LOGIN_URL = PREPROD_URL
        print(f"Environment set to: {self.environment}")
        print(f"Login URL: {LOGIN_URL}")
    
    def automation_sleep(self, seconds):
        """Sleep for specified seconds divided by automation speed
        - 2x = 2x faster (half the wait time)
        - 1x = normal speed
        - 0.25x = 4x slower (4 times the wait time)
        """
        import time
        time.sleep(seconds / self.automation_speed)
    
    def apply_theme_colors(self):
        if self.is_dark_mode:
            self.bg_dark = "#0f0f0f"
            self.bg_card = "#1a1a1a"
            self.bg_light = "#252525"
            self.accent = "#0052CC"
            self.text_light = "#ffffff"
            self.text_dim = "#a0a0a0"
        else:
            self.bg_dark = "#f5f3f0"
            self.bg_card = "#ffffff"
            self.bg_light = "#fafaf8"
            self.accent = "#0066ff"
            self.text_light = "#1a1a1a"
            self.text_dim = "#000000"  # Black text for better readability in light mode
    
    def enable_dark_title_bar(self, window):
        """Enable dark mode title bar on Windows 11"""
        try:
            import ctypes
            window.update()
            hwnd = ctypes.windll.user32.GetParent(window.winfo_id())
            # DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            rendering_policy = 20
            value = ctypes.c_int(2 if self.is_dark_mode else 0)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, rendering_policy, ctypes.byref(value), ctypes.sizeof(value)
            )
        except:
            pass  # Silently fail if not supported
    
    def style_dialog(self, dialog):
        """Apply dark mode styling to dialog window"""
        if self.is_dark_mode:
            dialog.configure(bg="#1a1a1a")
            self.enable_dark_title_bar(dialog)
            # Style all widgets in dialog
            for widget in dialog.winfo_children():
                try:
                    widget.configure(bg="#1a1a1a", fg="#ffffff")
                except:
                    pass
    
    def clear_container(self):
        if self.current_screen:
            try:
                self.current_screen.destroy()
            except:
                pass
        for widget in self.container.winfo_children():
            widget.destroy()
    
    def get_chrome_options(self):
        """Get Chrome options with headless mode setting applied"""
        from selenium.webdriver.chrome.options import Options
        chrome_options = Options()
        
        # Always use incognito mode to avoid cookies and prompts
        chrome_options.add_argument("--incognito")
        
        if self.headless_mode:
            chrome_options.add_argument("--headless=new")
        else:
            # Set window to 3/4 screen size when visible
            chrome_options.add_argument("--window-size=1440,810")  # 3/4 of 1920x1080
            chrome_options.add_argument("--window-position=0,0")
        
        # Standard options
        chrome_options.add_argument("--force-device-scale-factor=0.75")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        
        return chrome_options
    
    def show_login_screen(self):
        self.update_login_url()  # Set correct URL based on environment
        self.clear_container()
        self.current_screen = LoginScreen(self.container, self)
        self.root.title("SMART Assistant - Login")
    
    def show_main_menu(self):
        self.clear_container()
        self.current_screen = MainMenuScreen(self.container, self)
        self.root.title("SMART Assistant - Main Menu")
    
    def show_add_users_menu(self):
        self.clear_container()
        self.current_screen = AddUsersMenuScreen(self.container, self)
        self.root.title("SMART Assistant - Add Users")
    
    def show_standards_scans_menu(self):
        self.clear_container()
        self.current_screen = StandardsScansMenuScreen(self.container, self)
        self.root.title("SMART Assistant - Standards Scans")
    

    def show_user_deactivation(self):
        self.clear_container()
        placeholder = tk.Frame(self.container, bg=self.bg_dark)
        placeholder.pack(fill="both", expand=True)
        
        # Header
        header_frame = tk.Frame(placeholder, bg=self.bg_card, height=80)
        header_frame.pack(fill="x", pady=0)
        header_frame.pack_propagate(False)
        
        # Back button in top right
        tk.Button(header_frame, 
                  text="← Back to Menu", 
                  command=self.show_main_menu,
                  bg=self.bg_light, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(relx=1.0, x=-20, y=20, anchor="ne")
        
        tk.Label(header_frame, 
                 text="User Deactivation Manager", 
                 font=("Segoe UI", 18, "bold"),
                 bg=self.bg_card,
                 fg=self.accent).pack(pady=20)
        
        # Main content area
        content_frame = tk.Frame(placeholder, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Left side: Console Log
        log_frame = tk.Frame(content_frame, bg=self.bg_card, relief="solid", borderwidth=1)
        log_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        tk.Label(log_frame, text="Console Log", font=("Segoe UI", 11, "bold"), 
                 bg=self.bg_light if self.is_dark_mode else "#fafaf8", 
                 fg=self.accent, anchor="w", padx=10, pady=8).pack(fill="x")
        
        self.log_text = scrolledtext.ScrolledText(log_frame, 
                                             wrap=tk.WORD, 
                                             font=("Consolas", 9),
                                             bg=self.bg_card,
                                             fg=self.text_light,
                                             relief="flat",
                                             state="disabled")
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Right side: Stats panel
        stats_panel = tk.Frame(content_frame, bg=self.bg_card, relief="solid", borderwidth=1, width=300)
        stats_panel.pack(side="right", fill="y")
        stats_panel.pack_propagate(False)
        
        # Header with Stop Scan link
        header_frame = tk.Frame(stats_panel, bg=self.bg_light if self.is_dark_mode else "#fafaf8")
        header_frame.pack(fill="x")
        
        tk.Label(header_frame, text="Scan Statistics", font=("Segoe UI", 11, "bold"), 
                 bg=self.bg_light if self.is_dark_mode else "#fafaf8", 
                 fg=self.accent, anchor="w", padx=10, pady=8).pack(side="left")
        
        # Stop scan link on the right
        self.stop_scan_link = tk.Label(header_frame, text="⏹ Stop", 
                                       font=("Segoe UI", 9, "bold"),
                                       bg=self.bg_light if self.is_dark_mode else "#fafaf8",
                                       fg="#ff4444", cursor="hand2", padx=10)
        self.stop_scan_link.pack(side="right", pady=8)
        self.stop_scan_link.bind("<Button-1>", lambda e: self.stop_scan())
        self.stop_scan_link.pack_forget()  # Hidden by default
        
        stats_content = tk.Frame(stats_panel, bg=self.bg_card)
        stats_content.pack(fill="both", expand=True, padx=15, pady=10)  # Reduced pady from 15 to 10
        
        # Stats labels - Reduced spacing - Store references for live updates
        tk.Label(stats_content, text="Current Page:", font=("Segoe UI", 9), 
                 bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        self.stat_page_label = tk.Label(stats_content, text="0", font=("Segoe UI", 16, "bold"),
                 bg=self.bg_card, fg=self.accent, anchor="w")
        self.stat_page_label.pack(fill="x", pady=(0, 8))
        
        tk.Label(stats_content, text="Users Found:", font=("Segoe UI", 9), 
                 bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        self.stat_found_label = tk.Label(stats_content, text="0", font=("Segoe UI", 16, "bold"),
                 bg=self.bg_card, fg="#4CAF50", anchor="w")  # Green like main menu
        self.stat_found_label.pack(fill="x", pady=(0, 8))
        
        tk.Label(stats_content, text="Users Skipped:", font=("Segoe UI", 9), 
                 bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        self.stat_skipped_label = tk.Label(stats_content, text="0", font=("Segoe UI", 16, "bold"),
                 bg=self.bg_card, fg="#ffaa00", anchor="w")
        self.stat_skipped_label.pack(fill="x", pady=(0, 8))
        
        tk.Label(stats_content, text="Status:", font=("Segoe UI", 9), 
                 bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        self.status_label = tk.Label(stats_content, text="Ready", font=("Segoe UI", 12, "bold"),
                 bg=self.bg_card, fg="#ff4444", anchor="w")  # Red for visibility
        self.status_label.pack(fill="x", pady=(0, 8))
        
        # Progress bar
        tk.Label(stats_content, text="Progress:", font=("Segoe UI", 9), 
                 bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        
        progress_frame = tk.Frame(stats_content, bg=self.bg_light, height=25)
        progress_frame.pack(fill="x", pady=(0, 5))
        progress_frame.pack_propagate(False)
        
        self.progress_bar = tk.Frame(progress_frame, bg=self.accent, height=25)
        self.progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        self.progress_percent_label = tk.Label(stats_content, text="0%", font=("Segoe UI", 9),
                 bg=self.bg_card, fg=self.text_dim, anchor="center")
        self.progress_percent_label.pack(fill="x")
        
        # Buttons - Darker colors in dark mode
        button_color = "#0052CC" if self.is_dark_mode else "#0066ff"
        button_color_dark = "#0052CC" if self.is_dark_mode else "#0066ff"  # 20% darker for dark mode
        button_text_color = "black"  # Black text for all buttons for readability
        
        self.start_scan_button = tk.Button(stats_content, 
                  text="▶ Start Scan", 
                  command=self.handle_start_scan,
                  bg=button_color_dark, 
                  fg=button_text_color,
                  font=("Segoe UI", 11, "bold"),
                  relief="flat",
                  cursor="hand2",
                  pady=8,
                  highlightthickness=2,
                  highlightbackground="#333333")
        self.start_scan_button.pack(fill="x", pady=(15, 0))
        
        tk.Button(stats_content, 
                  text="📊 Deactivate from Spreadsheet", 
                  command=self.handle_deactivate_from_spreadsheet,
                  bg=button_color_dark, 
                  fg=button_text_color,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  pady=6,
                  highlightthickness=2,
                  highlightbackground="#333333").pack(fill="x", pady=(8, 0))
        
        # Deactivate Users Button (renamed from Add Multiple Users)
        self.batch_deactivate_button = tk.Button(stats_content, 
                  text="👥 Deactivate Users", 
                  command=lambda: self.handle_batch_deactivate(),
                  bg=button_color_dark, 
                  fg=button_text_color,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  pady=6,
                  highlightthickness=2,
                  highlightbackground="#333333")
        self.batch_deactivate_button.pack(fill="x", pady=(8, 0))
        
        
        # Bind click event
        
        # Flag to track if scan should stop
        self.should_stop_scan = False
        
        self.current_screen = placeholder
        self.root.title("SMART Assistant - User Deactivation")
    
    def stop_scan(self):
        """Stop the current scan immediately"""
        self.should_stop_scan = True
        self.log_to_ui("\n⏹ STOP REQUESTED - Stopping scan immediately...\n")
        self.stop_scan_link.config(text="Stopping...", fg="#999999")
        self.update_progress(0, "Stopping scan...")
        
        # Try to close Chrome immediately if it's open
        try:
            if hasattr(self, 'current_driver') and self.current_driver:
                self.current_driver.quit()
        except:
            pass
    
    
    def handle_batch_deactivate(self):
        """Handle batch deactivation of multiple users"""
        # Open the batch deactivation dialog
        self.show_batch_deactivation_dialog()
    
    def show_batch_deactivation_dialog(self):
        """Show dialog for adding multiple users to deactivate"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Multiple Users for Deactivation")
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        # Center dialog
        dialog.geometry("600x500")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Header
        tk.Label(dialog, text="Add Multiple Users for Deactivation", 
                font=("Segoe UI", 14, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.accent).pack(pady=15, padx=20)
        
        # Instructions
        tk.Label(dialog, text="Add username and ticket number pairs. All users will be deactivated in sequence.", 
                font=("Segoe UI", 9),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_dim,
                wraplength=550).pack(pady=(0, 10), padx=20)
        
        # User list storage
        user_list = []
        
        # List display frame
        list_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Scrollable list
        canvas = tk.Canvas(list_frame, bg=self.bg_light if self.is_dark_mode else "#f5f5f5", 
                          highlightthickness=0, height=200)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.bg_light if self.is_dark_mode else "#f5f5f5")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Header row for the list
        header_frame = tk.Frame(scrollable_frame, bg=self.bg_light if self.is_dark_mode else "#f5f5f5")
        header_frame.pack(fill="x", padx=5, pady=5)
        tk.Label(header_frame, text="Username", font=("Segoe UI", 9, "bold"),
                bg=self.bg_light if self.is_dark_mode else "#f5f5f5",
                fg=self.text_light, width=20, anchor="w").pack(side="left", padx=5)
        tk.Label(header_frame, text="Ticket #", font=("Segoe UI", 9, "bold"),
                bg=self.bg_light if self.is_dark_mode else "#f5f5f5",
                fg=self.text_light, width=15, anchor="w").pack(side="left", padx=5)
        
        # Count label reference (created before functions need it)
        count_label_var = tk.StringVar(value="Users added: 0")
        
        def refresh_list():
            """Refresh the display of users"""
            # Clear existing items (except header)
            for widget in scrollable_frame.winfo_children()[1:]:
                widget.destroy()
            
            # Add each user to the list
            for idx, (username, ticket) in enumerate(user_list):
                item_frame = tk.Frame(scrollable_frame, bg=self.bg_card if self.is_dark_mode else "#ffffff",
                                     relief="solid", borderwidth=1)
                item_frame.pack(fill="x", padx=5, pady=2)
                
                tk.Label(item_frame, text=username, font=("Segoe UI", 10),
                        bg=self.bg_card if self.is_dark_mode else "#ffffff",
                        fg=self.text_light, width=20, anchor="w").pack(side="left", padx=5, pady=5)
                tk.Label(item_frame, text=ticket, font=("Segoe UI", 10),
                        bg=self.bg_card if self.is_dark_mode else "#ffffff",
                        fg=self.text_light, width=15, anchor="w").pack(side="left", padx=5, pady=5)
                
                # Delete button
                tk.Button(item_frame, text="✕", 
                         command=lambda i=idx: remove_user(i),
                         bg="#ff4444", fg="white",
                         font=("Segoe UI", 8, "bold"),
                         relief="flat", cursor="hand2",
                         width=3).pack(side="right", padx=5, pady=5)
            
            # Update count label
            count_label_var.set(f"Users added: {len(user_list)}")
        
        def remove_user(index):
            """Remove a user from the list"""
            if 0 <= index < len(user_list):
                user_list.pop(index)
                refresh_list()
        
        def add_user():
            """Add a user to the list"""
            # Get username
            username = self.show_dark_input("Username", "Enter username:")
            if not username:
                return
            
            # Get ticket
            ticket = self.show_dark_input("Ticket Number", "Enter ticket number:")
            if not ticket:
                return
            
            # Add to list
            user_list.append((username, ticket))
            refresh_list()
        
        # Add User button
        add_btn_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        add_btn_frame.pack(fill="x", padx=20, pady=(10, 5))
        
        tk.Button(add_btn_frame, text="+ Add User", command=add_user,
                 bg=self.accent, fg="white" if not self.is_dark_mode else "black",
                 font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                 width=15).pack(side="left", padx=5)
        
        tk.Label(add_btn_frame, textvariable=count_label_var, 
                font=("Segoe UI", 9),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_dim).pack(side="left", padx=10)
        
        # Action buttons
        btn_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        btn_frame.pack(pady=15)
        
        def start_batch():
            """Start batch deactivation"""
            if not user_list:
                self.show_dark_messagebox("No Users", "Please add at least one user to deactivate.", "warning")
                return
            
            # Confirm
            if not self.show_dark_confirm("Confirm Batch Deactivation", 
                                         f"Deactivate {len(user_list)} user(s)?"):
                return
            
            dialog.destroy()
            
            # Run batch deactivation
            self.run_batch_deactivation(user_list)
        
        tk.Button(btn_frame, text="Start Deactivation", command=start_batch, width=18,
                 bg="#4CAF50", fg="white",
                 font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=10,
                 bg=self.bg_light, fg=self.text_light,
                 font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=5)
        
        dialog.wait_window()
    
    def run_batch_deactivation(self, user_list):
        """Run batch deactivation for multiple users"""
        import threading
        
        # Disable both buttons during operation
        self.batch_deactivate_button.config(state="disabled", text="Processing...")
        
        def run_batch():
            driver = None
            success_count = 0
            fail_count = 0
            
            try:
                self.log_to_ui("="*60)
                self.log_to_ui(f"BATCH DEACTIVATION: {len(user_list)} users")
                self.log_to_ui("="*60 + "\n")
                self.update_progress(5, "Launching Chrome...")
                
                # Create driver
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service
                
                chrome_options = self.get_chrome_options()
                
                if self.headless_mode:
                    chrome_options.add_argument("--window-size=1920,1080")
                    chrome_options.add_argument("--start-maximized")
                else:
                    chrome_options.add_argument("--start-maximized")
                
                service = Service()
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.set_page_load_timeout(20)
                
                self.update_progress(10, "Logging in...")
                self.log_to_ui("→ Logging into SMART system...")
                
                # Login sequence
                driver.get(get_url("/auth/login", self))
                
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.by import By
                import time
                
                actions = ActionChains(driver)
                wait = WebDriverWait(driver, 10)
                
                self.automation_sleep(1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.username).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.password).perform()
                self.automation_sleep(0.2)
                
                login_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
                ))
                driver.execute_script("arguments[0].click();", login_button)
                self.automation_sleep(2)
                
                self.update_progress(20, "Selecting program...")
                self.log_to_ui("→ Selecting program...")
                
                # Select program
                try:
                    self.automation_sleep(1)
                    actions.send_keys(Keys.TAB).perform()
                    self.automation_sleep(0.1)
                    actions.send_keys("0000").perform()
                    self.automation_sleep(0.3)
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.1)
                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(0.2)
                    
                    try:
                        submit_button = wait.until(EC.element_to_be_clickable(
                            (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
                        ))
                        driver.execute_script("arguments[0].click();", submit_button)
                        self.automation_sleep(1)
                        self.log_to_ui("  ✓ Program selected and submitted")
                    except:
                        self.log_to_ui("  ! Could not find SUBMIT button")
                    
                    self.automation_sleep(2)
                    
                except Exception as e:
                    self.log_to_ui(f"  ! Program selection: {str(e)}")
                
                self.update_progress(25, "Logged in successfully")
                self.log_to_ui("  ✓ Login complete\n")
                
                # Process each user
                for idx, (username, ticket) in enumerate(user_list, 1):
                    progress_base = 25 + int((idx - 1) / len(user_list) * 70)
                    progress_step = int(70 / len(user_list))
                    
                    self.log_to_ui(f"\n[{idx}/{len(user_list)}] Processing: {username} (Ticket {ticket})")
                    self.update_progress(progress_base, f"User {idx}/{len(user_list)}: {username}")
                    
                    # Deactivate user
                    success = deactivate_user_by_ticket(driver, username, ticket, self.automation_speed, self)
                    
                    if success:
                        success_count += 1
                        self.log_to_ui(f"  ✓ User '{username}' deactivated successfully!")
                    else:
                        fail_count += 1
                        self.log_to_ui(f"  ✗ Failed to deactivate '{username}'")
                    
                    self.automation_sleep(0.5)  # Brief pause between users
                
                self.update_progress(100, "Complete!")
                
                # Summary
                self.log_to_ui("\n" + "="*60)
                self.log_to_ui("BATCH DEACTIVATION COMPLETE")
                self.log_to_ui(f"Total users: {len(user_list)}")
                self.log_to_ui(f"✓ Successful: {success_count}")
                self.log_to_ui(f"✗ Failed: {fail_count}")
                self.log_to_ui("="*60 + "\n")
                
                # Show summary dialog
                self.show_dark_messagebox("Batch Complete", 
                                        f"Processed {len(user_list)} users\n✓ Success: {success_count}\n✗ Failed: {fail_count}",
                                        "info")
                
            except Exception as e:
                self.log_to_ui(f"\n✗ BATCH ERROR: {str(e)}\n")
                print(f"Error: {str(e)}")
                self.show_dark_messagebox("Error", f"Batch error: {str(e)}", "error")
            finally:
                # Close driver
                if driver:
                    try:
                        self.log_to_ui("→ Closing browser...")
                        driver.quit()
                        self.log_to_ui("  ✓ Browser closed\n")
                    except:
                        pass
                
                # Re-enable buttons
                self.update_progress(0, "Ready")
                self.batch_deactivate_button.config(state="normal", text="👥 Deactivate Users")
        
        threading.Thread(target=run_batch, daemon=True).start()
    
    def handle_deactivate_from_spreadsheet(self):
        """Handle deactivate from spreadsheet button - parse Excel/CSV and deactivate users"""
        from tkinter import filedialog
        import re
        import csv
        
        # Prompt user to select file
        file_path = filedialog.askopenfilename(
            title="Select Spreadsheet",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            self.log_to_ui("="*60)
            self.log_to_ui(f"LOADING SPREADSHEET")
            self.log_to_ui("="*60 + "\n")
            self.log_to_ui(f"→ File: {file_path}\n")
            
            # Read spreadsheet
            rows = []
            if file_path.endswith('.csv'):
                # Read CSV
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
            else:
                # Read Excel
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = [[cell.value if cell.value is not None else '' for cell in row] for row in ws.rows]
                    wb.close()
                except ImportError:
                    self.show_dark_messagebox("Missing Library",
                                            "openpyxl is required to read Excel files.\n\n"
                                            "Install it with:\npip install openpyxl\n\n"
                                            "Or use a CSV file instead.",
                                            "error")
                    return
            
            self.log_to_ui(f"✓ Loaded {len(rows)} rows\n")
            
            # Try to find header row and column indices
            name_col = None
            username_col = None
            first_name_col = None
            last_name_col = None
            
            for i, row in enumerate(rows[:5]):  # Check first 5 rows for headers
                for j, cell in enumerate(row):
                    cell_lower = str(cell).lower().strip()
                    if 'username' in cell_lower or 'user name' in cell_lower:
                        username_col = j
                    elif 'first' in cell_lower and 'name' in cell_lower:
                        first_name_col = j
                    elif 'last' in cell_lower and 'name' in cell_lower:
                        last_name_col = j
                    elif cell_lower in ['name', 'full name', 'employee name', 'employee']:
                        name_col = j
            
            self.log_to_ui(f"  Column detection: username={username_col}, name={name_col}, first={first_name_col}, last={last_name_col}\n")
            
            # Parse users from spreadsheet
            users = []
            username_pattern = re.compile(r'^\d{4}[a-zA-Z][a-zA-Z]+$')  # 0000FirstnameLastname
            seen_usernames = set()
            
            for row_idx, row in enumerate(rows):
                if row_idx < 1 and (name_col is not None or username_col is not None):
                    continue  # Skip header row if we found column headers
                
                username = None
                name = None
                
                # Try to get username from detected column first
                if username_col is not None and username_col < len(row):
                    potential = str(row[username_col]).strip()
                    clean = re.sub(r'[^\w]', '', potential)
                    if username_pattern.match(clean):
                        username = clean
                
                # If no username column, search entire row
                if not username:
                    for cell in row:
                        clean_word = re.sub(r'[^\w]', '', str(cell))
                        if username_pattern.match(clean_word) and clean_word not in seen_usernames:
                            username = clean_word
                            break
                
                if username and username not in seen_usernames:
                    seen_usernames.add(username)
                    
                    # Try to get name from detected columns
                    if first_name_col is not None and last_name_col is not None:
                        if first_name_col < len(row) and last_name_col < len(row):
                            first = str(row[first_name_col]).strip()
                            last = str(row[last_name_col]).strip()
                            if first and last:
                                name = f"{first} {last}"
                    elif name_col is not None and name_col < len(row):
                        name = str(row[name_col]).strip()
                    
                    # If no name from columns, try to find capitalized words
                    if not name:
                        row_str = ' '.join(str(cell) for cell in row if cell)
                        words = row_str.split()
                        name_candidates = [w for w in words if len(w) > 2 and w[0].isupper() and not w.isdigit() and not username_pattern.match(w)]
                        if len(name_candidates) >= 2:
                            name = ' '.join(name_candidates[:2])
                        elif len(name_candidates) == 1:
                            name = name_candidates[0]
                    
                    users.append({
                        'username': username,
                        'name': name or username,
                        'ticket': ''
                    })
            
            self.log_to_ui(f"✓ Found {len(users)} users with valid usernames\n")
            
            if len(users) == 0:
                self.show_dark_messagebox("No Users Found",
                                        "No valid usernames found in spreadsheet.\n\n"
                                        "Expected format: 0000FirstnameLastname\n"
                                        "(4 digits followed by first initial and last name, no spaces)\n\n"
                                        "Examples: 0000JSmith, 1234MDoe",
                                        "warning")
                return
            
            # Show selection dialog with ticket input
            self.show_spreadsheet_selection_dialog(users)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.log_to_ui(f"\n✗ ERROR: {str(e)}\n")
            self.log_to_ui(f"{error_details}\n")
            self.show_dark_messagebox("Error", f"Failed to read spreadsheet:\n\n{str(e)}", "error")
    
    def handle_start_scan(self):
        """Handle start scan button - scans for inactive users"""
        import threading
        
        # Show filter management dialog
        if not self.show_filter_management_dialog():
            # User clicked Back
            return
        
        logger.info("="*60)
        logger.info(f"USER SCAN STARTED - Inactive {INACTIVE_DAYS}+ days")
        logger.info(f"Settings: Headless={self.headless_mode}, Speed={self.automation_speed}, Environment={self.environment}")
        logger.info(f"Custom filters: {self.custom_filters}")
        logger.info("="*60)
        
        # Disable buttons during scan
        self.start_scan_button.config(state="disabled", text="Scanning...")
        self.batch_deactivate_button.config(state="disabled")
        self.stop_scan_link.pack(side="right", pady=8)
        self.stop_scan_link.config(text="⏹ Stop", fg="#ff4444")
        self.should_stop_scan = False  # Reset stop flag
        
        # Show warning if not in headless mode
        if not self.headless_mode:
            self.show_dark_messagebox("Non-Headless Mode Active",
                                    "⚠️ Chrome is running in visible mode.\n\n"
                                    "Please DO NOT:\n"
                                    "• Minimize the Chrome window\n"
                                    "• Click away from Chrome\n"
                                    "• Close Chrome manually\n\n"
                                    "Keep Chrome visible so you can see the scan statistics.",
                                    "warning")
        
        def run_scan():
            driver = None
            candidates = []
            
            try:
                self.log_to_ui("="*60)
                self.log_to_ui(f"SCANNING FOR INACTIVE USERS ({INACTIVE_DAYS}+ DAYS)")
                self.log_to_ui("="*60 + "\n")
                self.update_progress(5, "Launching Chrome...")
                
                # Create driver
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service
                
                chrome_options = self.get_chrome_options()
                
                if self.headless_mode:
                    chrome_options.add_argument("--window-size=1920,1080")
                    chrome_options.add_argument("--start-maximized")
                else:
                    # Resize Chrome to 60% width and position on left side
                    chrome_options.add_argument("--window-size=1152,1080")  # 60% of 1920
                    chrome_options.add_argument("--window-position=0,0")
                
                service = Service()
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.set_page_load_timeout(20)
                
                # Store driver reference for immediate stop capability
                self.current_driver = driver
                
                self.update_progress(10, "Logging in...")
                self.log_to_ui("→ Logging into SMART system...")
                
                # Login sequence
                driver.get(LOGIN_URL)
                
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.by import By
                import time
                
                actions = ActionChains(driver)
                wait = WebDriverWait(driver, 10)
                
                self.automation_sleep(1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.username).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.password).perform()
                self.automation_sleep(0.2)
                
                login_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
                ))
                driver.execute_script("arguments[0].click();", login_button)
                self.automation_sleep(2)
                
                self.update_progress(20, "Selecting program...")
                self.log_to_ui("→ Selecting program...")
                
                # Select program
                try:
                    self.automation_sleep(1)
                    actions.send_keys(Keys.TAB).perform()
                    self.automation_sleep(0.1)
                    actions.send_keys("0000").perform()
                    self.automation_sleep(0.3)
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.1)
                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(0.2)
                    
                    try:
                        submit_button = wait.until(EC.element_to_be_clickable(
                            (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
                        ))
                        driver.execute_script("arguments[0].click();", submit_button)
                        self.automation_sleep(1)
                        self.log_to_ui("  ✓ Program selected")
                    except:
                        self.log_to_ui("  ! Could not find SUBMIT button")
                    
                    self.automation_sleep(2)
                    
                except Exception as e:
                    self.log_to_ui(f"  ! Program selection: {str(e)}")
                
                self.update_progress(30, "Navigating to users page...")
                self.log_to_ui("→ Navigating to users page...")
                
                # Navigate to users page
                driver.get(get_url("/main/users", self))
                self.automation_sleep(1)
                driver.find_element(By.TAG_NAME, "body").click()
                self.automation_sleep(0.3)
                
                self.log_to_ui("  ✓ On users page")
                
                # Sort by Last Login (descending) - click the sort arrow
                self.update_progress(33, "Sorting by Last Login...")
                self.log_to_ui("→ Sorting by Last Login (descending)...")
                
                try:
                    # Scroll to top first
                    driver.execute_script("window.scrollTo(0, 0);")
                    self.automation_sleep(0.5)
                    
                    # Method 1: Try clicking the path element directly (most reliable from beta)
                    result = driver.execute_script("""
                        var path = document.querySelector("#root > div.collapsed.AdminLayout_admin_layout_container__8z6D1 > div.AdminLayout_admin_body_container__qHWPk > div > div.content-container > div > div.maintable-content.table-responsive > div.table-scroll > table > thead > tr > th:nth-child(7) > div > svg > path");
                        if (path) {
                            var event = new MouseEvent('click', {
                                view: window,
                                bubbles: true,
                                cancelable: true
                            });
                            path.dispatchEvent(event);
                            return true;
                        }
                        return false;
                    """)
                    
                    if result:
                        self.automation_sleep(1.0)
                        self.log_to_ui("  ✓ Sorted by Last Login (descending)")
                    else:
                        # Fallback: Try finding the header by text
                        headers = driver.find_elements(By.TAG_NAME, "th")
                        for header in headers:
                            if "last login" in header.text.lower():
                                driver.execute_script("arguments[0].click();", header)
                                self.automation_sleep(0.5)
                                # Click again for descending
                                driver.execute_script("arguments[0].click();", header)
                                self.automation_sleep(0.5)
                                self.log_to_ui("  ✓ Sorted by Last Login (descending)")
                                break
                        else:
                            self.log_to_ui("  ! Could not sort - continuing anyway")
                        
                except Exception as e:
                    self.log_to_ui(f"  ! Sorting failed: {str(e)} - continuing anyway")
                
                self.log_to_ui("")
                
                # Set page size to 100 - CORRECT NAVIGATION FROM USER
                self.update_progress(36, "Setting page size to 100...")
                self.log_to_ui("→ Setting page size to 100 users per page...")
                
                try:
                    # Wait a moment after sorting
                    self.automation_sleep(0.5)
                    
                    # Shift+Tab 14 times to reach the page size dropdown
                    self.log_to_ui("  → Navigating to page size dropdown (Shift+Tab 14x)...")
                    for i in range(14):
                        actions.key_down(Keys.SHIFT).send_keys(Keys.TAB).key_up(Keys.SHIFT).perform()
                        self.automation_sleep(0.15)  # Increased from 0.05 - visible delay between tabs
                    
                    self.automation_sleep(0.3)  # Increased from 0.2
                    
                    # Space to open dropdown
                    self.log_to_ui("  → Opening dropdown (Space)...")
                    actions.send_keys(Keys.SPACE).perform()
                    self.automation_sleep(0.3)  # Increased from 0.2
                    
                    # Down arrow twice to select 100
                    self.log_to_ui("  → Selecting 100 (Down Down)...")
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.3)  # Increased from 0.1
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.3)  # Increased from 0.1
                    
                    # Enter to confirm
                    self.log_to_ui("  → Confirming selection (Enter)...")
                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(1.5)  # Wait for page to reload with 100 items
                    
                    self.log_to_ui("  ✓ Page size set to 100")
                    
                except Exception as e:
                    self.log_to_ui(f"  ! Page size change failed: {str(e)}")
                    self.log_to_ui("  ! Continuing with default page size")
                
                self.log_to_ui("")
                self.update_progress(40, "Starting scan...")
                
                # Run the scan
                candidates = self.scan_for_inactive_users(driver, actions, wait)
                
                self.update_progress(100, "Scan complete!")
                
                if len(candidates) > 0:
                    self.log_to_ui(f"\n✓ Found {len(candidates)} inactive users\n")
                    
                    # Minimize Chrome if in non-headless mode so dialog is visible
                    if not self.headless_mode and driver:
                        try:
                            driver.minimize_window()
                            self.log_to_ui("→ Chrome minimized\n")
                        except:
                            pass
                    
                    self.show_user_selection_dialog(candidates, driver)
                else:
                    self.log_to_ui("\n✓ No inactive users found OR scan paused\n")
                    # Close browser if no users found
                    if driver:
                        try:
                            self.log_to_ui("→ Closing browser...")
                            driver.quit()
                            self.log_to_ui("  ✓ Browser closed\n")
                        except:
                            pass
                
            except Exception as e:
                self.log_to_ui(f"\n✗ SCAN ERROR: {str(e)}\n")
                print(f"Error: {str(e)}")
                self.show_dark_messagebox("Error", f"Scan error: {str(e)}", "error")
                # Close browser on error
                if driver:
                    try:
                        driver.quit()
                    except:
                        pass
            finally:
                # Only close browser if we actually finished scanning (not paused)
                # For now, don't close it automatically - let user close manually
                pass
                
                # Re-enable buttons only if no candidates (otherwise keep disabled until deactivation done)
                if len(candidates) == 0:
                    self.update_progress(0, "Ready")
                    self.start_scan_button.config(state="normal", text="▶ Start Scan")
                    self.batch_deactivate_button.config(state="normal")
                
                # Always disable stop button when scan ends
                self.stop_scan_link.pack_forget()
                
                # Clear driver reference
                self.current_driver = None
        
        threading.Thread(target=run_scan, daemon=True).start()
    
    def scan_for_inactive_users(self, driver, actions, wait):
        """Scan users page for inactive users - EXACT COPY FROM WORKING BETA"""
        import time
        scan_start_time = time.time()
        
        candidates = []
        skipped = 0
        page = 1
        previous_page_hash = None
        early_stop = False
        
        # Use historical data for better estimates
        if hasattr(self, 'last_scan_pages') and self.last_scan_pages:
            estimated_pages = self.last_scan_pages
        else:
            estimated_pages = 45
        
        self.log_to_ui(f"→ Scanning for users inactive {INACTIVE_DAYS}+ days...")
        self.log_to_ui(f"  Estimated pages: {estimated_pages} (based on previous scan)\n")
        
        while True:
            # Check if stop was requested
            if self.should_stop_scan:
                self.log_to_ui(f"\n⏹ Scan stopped by user at page {page}\n")
                break
            
            # Better progress calculation using historical data
            if hasattr(self, 'last_scan_time') and self.last_scan_time and estimated_pages:
                # Time-based progress
                elapsed = time.time() - scan_start_time
                expected_total_time = self.last_scan_time
                time_progress = min((elapsed / expected_total_time) * 55, 55)
                
                # Page-based progress
                page_progress = min((page / estimated_pages) * 55, 55)
                
                # Use the average for smoother progress
                progress = 40 + int((time_progress + page_progress) / 2)
            else:
                # Fallback to page-based only
                progress = 40 + min((page / estimated_pages) * 55, 55)
            
            self.update_progress(int(progress), f"Scanning page {page}...")
            
            try:
                # Find table once per page with retry logic
                max_retries = 3
                table = None
                for retry in range(max_retries):
                    try:
                        table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, TABLE_SELECTOR)))
                        break
                    except:
                        if retry < max_retries - 1:
                            self.log_to_ui(f"  Retry {retry + 1}/3 loading page...")
                            self.automation_sleep(1)
                            # Try refreshing the page state
                            try:
                                driver.execute_script("return document.readyState")
                            except:
                                self.log_to_ui(f"  Browser connection lost, stopping scan")
                                break
                        else:
                            raise
                
                if not table:
                    self.log_to_ui(f"\nCould not load page {page}, stopping scan")
                    break
                
                # Get all rows at once - much faster than finding one by one
                rows = table.find_elements(By.CSS_SELECTOR, ROWS_SELECTOR)

                # Quick duplicate check using first 3 rows
                page_hash = hash(tuple(row.text for row in rows[:3]))
                if previous_page_hash == page_hash and page > 1:
                    self.log_to_ui(f"Page {page}: Duplicate detected - stopping\n")
                    break
                previous_page_hash = page_hash

                page_found = 0
                
                # Process all rows in batch
                for row in rows:
                    try:
                        # Get all cells at once instead of one by one
                        cells = row.find_elements(By.TAG_NAME, "td")
                        
                        # Quick check for minimum columns
                        if len(cells) <= max(COL_LAST_LOGIN, COL_USERNAME, COL_JOB_TITLE, COL_NAME_LINK):
                            continue

                        # Get all text values at once - cache them
                        name_text = cells[COL_NAME_LINK].text
                        job_title = cells[COL_JOB_TITLE].text
                        username = cells[COL_USERNAME].text
                        last_login_str = cells[COL_LAST_LOGIN].text
                        
                        # Comprehensive string checks - catch all variations
                        name_lower = name_text.lower()
                        
                        # Skip deactivated users (all variations)
                        deactivated_patterns = ["deactivat", "deact", "(deact", "- deact", "end access"]
                        if any(pattern in name_lower for pattern in deactivated_patterns):
                            skipped += 1
                            continue
                        
                        # Skip terminated users (all variations)
                        terminated_patterns = ["terminat", "termin", "(termin", "- termin", "(term", "- term"]
                        if any(pattern in name_lower for pattern in terminated_patterns):
                            skipped += 1
                            continue

                        job_lower = job_title.lower()
                        if "client" in job_lower:
                            skipped += 1
                            continue
                        
                        # Skip custom filtered values (checks Name, Username, and Job Title - exact match, case-insensitive)
                        username_lower = username.lower()
                        for custom_filter in self.custom_filters:
                            filter_lower = custom_filter.lower()
                            if (name_text.lower() == filter_lower or 
                                username_lower == filter_lower or 
                                job_title.lower() == filter_lower):
                                skipped += 1
                                break
                        else:
                            # Continue if no filter matched
                            pass
                        
                        # If we broke out of the loop (filter matched), skip this user
                        if any(name_text.lower() == f.lower() or username_lower == f.lower() or job_title.lower() == f.lower() 
                               for f in self.custom_filters):
                            continue

                        # Quick empty/invalid check
                        if not last_login_str or len(last_login_str) < 8:  # Valid dates are at least 8 chars
                            skipped += 1
                            continue
                            
                        login_lower = last_login_str.lower()
                        if login_lower in ["n/a", "na", "none", "-", "unknown"]:
                            skipped += 1
                            continue

                        # Parse date
                        try:
                            last_login_date = datetime.strptime(last_login_str, DATE_FORMAT)
                            threshold = datetime.now() - timedelta(days=INACTIVE_DAYS)
                            
                            # Check if too recent - early stop
                            if last_login_date >= threshold:
                                self.log_to_ui(f"\nPage {page}: User found with login < {INACTIVE_DAYS} days ago")
                                self.log_to_ui(f"  Name: {name_text}")
                                self.log_to_ui(f"  Username: {username}")
                                self.log_to_ui(f"  Last login: {last_login_str}")
                                self.log_to_ui("  → Stopping scan\n")
                                early_stop = True
                                break
                            
                            # Add to candidates - use cached values
                            display_name = name_text.strip() if name_text else "[No name]"
                            candidates.append((display_name, username.strip(), job_title.strip(), last_login_str.strip()))
                            page_found += 1
                            
                        except ValueError:
                            skipped += 1
                            continue

                    except Exception:
                        skipped += 1
                
                self.log_to_ui(f"Page {page}: Found {page_found} qualifying users (Total: {len(candidates)})")
                
                # Update statistics in UI
                self.stat_page_label.config(text=str(page))
                self.stat_found_label.config(text=str(len(candidates)))
                self.stat_skipped_label.config(text=str(skipped))
                self.root.update_idletasks()  # Force UI update
                
                if early_stop:
                    break

                # Navigate to next page - uses automation speed multiplier
                try:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    self.automation_sleep(0.3)  # Increased from 0.1

                    actions.send_keys(Keys.TAB * PAGINATION_TAB_COUNT).perform()
                    self.automation_sleep(0.4)  # Increased from 0.15

                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(0.3)  # Increased from 0.1

                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.3)  # Increased from 0.15

                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(1.0)  # Increased from 0.5 - wait for page to load

                    page += 1

                except Exception as e:
                    self.log_to_ui(f"\nNavigation failed: {str(e)}")
                    # Check if browser is still alive
                    try:
                        driver.current_url
                        self.log_to_ui("Attempting to continue...")
                        self.automation_sleep(1)
                        page += 1
                    except:
                        self.log_to_ui("Browser connection lost, stopping scan\n")
                        break

            except Exception as e:
                self.log_to_ui(f"\nScan error on page {page}: {str(e)}")
                # Check if it's a connection error
                try:
                    driver.current_url
                    self.log_to_ui("Attempting to recover...")
                    self.automation_sleep(2)
                except:
                    self.log_to_ui("Browser connection lost\n")
                    break
        
        # Save scan timing for future estimates
        scan_duration = time.time() - scan_start_time
        self.last_scan_time = scan_duration
        self.last_scan_pages = page
        self.save_settings()  # Persist the timing data
        
        self.log_to_ui(f"\n{'='*60}")
        self.log_to_ui("SCAN COMPLETE")
        self.log_to_ui(f"{'='*60}")
        self.log_to_ui(f"  Pages scanned:      {page}")
        self.log_to_ui(f"  Scan duration:      {int(scan_duration // 60)}m {int(scan_duration % 60)}s")
        self.log_to_ui(f"  Qualifying users:   {len(candidates)}")
        self.log_to_ui(f"  Users skipped:      {skipped}")
        if early_stop:
            self.log_to_ui(f"  Early stop:         Yes (found user < {INACTIVE_DAYS} days)")
        self.log_to_ui(f"{'='*60}\n")
        
        return candidates
    
    def show_user_selection_dialog(self, candidates, driver):
        """Show dialog to select which scanned users to deactivate"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Users to Deactivate")
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        # Make it large
        dialog.geometry("1200x700")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"1200x700+{x}+{y}")
        
        # Header
        tk.Label(dialog, text=f"{len(candidates)} Users Found", 
                font=("Segoe UI", 16, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.accent).pack(pady=15, padx=20)
        
        tk.Label(dialog, text=f"Select users to deactivate (inactive for {INACTIVE_DAYS}+ days)", 
                font=("Segoe UI", 10),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_dim).pack(pady=(0, 10), padx=20)
        
        # Control buttons
        control_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        control_frame.pack(fill="x", padx=20, pady=10)
        
        selected_count_var = tk.StringVar(value=f"✓ {len(candidates)} Selected")
        tk.Label(control_frame, textvariable=selected_count_var, 
                font=("Segoe UI", 11, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg="#4CAF50").pack(side="left", padx=10)
        
        # Treeview for user list
        list_frame = tk.Frame(dialog, bg=self.bg_light if self.is_dark_mode else "#f5f5f5",
                             relief="solid", borderwidth=1)
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Style treeview
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Scan.Treeview",
                       background=self.bg_card if self.is_dark_mode else "#ffffff",
                       foreground=self.text_light,
                       rowheight=28,
                       fieldbackground=self.bg_card if self.is_dark_mode else "#ffffff",
                       font=("Segoe UI", 9))
        style.configure("Scan.Treeview.Heading",
                       background=self.bg_light if self.is_dark_mode else "#e0e0e0",
                       foreground=self.accent,
                       font=("Segoe UI", 10, "bold"))
        style.map("Scan.Treeview",
                 background=[("selected", self.accent)],
                 foreground=[("selected", "white")])
        
        vsb = ttk.Scrollbar(list_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        
        tree = ttk.Treeview(list_frame,
                           columns=("name", "username", "jobtitle", "lastlogin"),
                           show="tree headings",
                           style="Scan.Treeview",
                           yscrollcommand=vsb.set)
        
        vsb.config(command=tree.yview)
        
        # Track sort state
        sort_reverse = {}
        
        def treeview_sort_column(col, reverse):
            """Sort treeview by column"""
            # Get all items
            items = [(tree.set(item, col), item) for item in tree.get_children('')]
            
            # Sort by column
            if col == "lastlogin":
                # Special handling for date column
                def parse_date(date_str):
                    try:
                        from datetime import datetime
                        return datetime.strptime(date_str, "%m/%d/%Y %I:%M %p")
                    except:
                        return datetime.min
                items.sort(key=lambda t: parse_date(t[0]), reverse=reverse)
            else:
                # String sort for other columns
                items.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)
            
            # Rearrange items
            for index, (val, item) in enumerate(items):
                tree.move(item, '', index)
            
            # Update header to show sort direction
            headers_text = {
                "name": "Name",
                "username": "Username", 
                "jobtitle": "Job Title",
                "lastlogin": "Last Login"
            }
            
            for c in headers_text:
                if c == col:
                    tree.heading(c, text=f"{headers_text[c]} {'▼' if reverse else '▲'}")
                else:
                    tree.heading(c, text=headers_text[c])
            
            # Toggle reverse for next click
            sort_reverse[col] = not reverse
        
        # Make headers clickable
        tree.heading("name", text="Name", command=lambda: treeview_sort_column("name", sort_reverse.get("name", False)))
        tree.heading("username", text="Username", command=lambda: treeview_sort_column("username", sort_reverse.get("username", False)))
        tree.heading("jobtitle", text="Job Title", command=lambda: treeview_sort_column("jobtitle", sort_reverse.get("jobtitle", False)))
        tree.heading("lastlogin", text="Last Login", command=lambda: treeview_sort_column("lastlogin", sort_reverse.get("lastlogin", False)))
        
        tree.heading("#0", text="☑", anchor="center")
        
        tree.column("#0", width=50, anchor="center")
        tree.column("name", width=250, anchor="w")
        tree.column("username", width=200, anchor="w")
        tree.column("jobtitle", width=200, anchor="w")
        tree.column("lastlogin", width=200, anchor="w")
        
        tree.pack(side="left", fill="both", expand=True)
        
        # Sort candidates by date (oldest first)
        sorted_candidates = sorted(candidates, key=lambda x: self._parse_date_safe(x[3]))
        
        # Add users to tree
        for name, username, job_title, last_login in sorted_candidates:
            display_title = job_title if job_title else "[No title]"
            tree.insert("", "end", 
                       text="☑",
                       values=(name, username, display_title, last_login))
        
        def update_count():
            count = sum(1 for item in tree.get_children() if tree.item(item, "text") == "☑")
            selected_count_var.set(f"✓ {count} Selected")
        
        def select_all():
            for item in tree.get_children():
                tree.item(item, text="☑")
            update_count()
        
        def deselect_all():
            for item in tree.get_children():
                tree.item(item, text="☐")
            update_count()
        
        def on_tree_click(event):
            region = tree.identify("region", event.x, event.y)
            if region == "tree" or region == "cell":
                item = tree.identify_row(event.y)
                if item:
                    current_text = tree.item(item, "text")
                    tree.item(item, text="☐" if current_text == "☑" else "☑")
                    update_count()
        
        tree.bind("<Button-1>", on_tree_click)
        
        tk.Button(control_frame, text="☑ Select All", command=select_all,
                 bg=self.accent, fg="black",
                 font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                 padx=12, pady=6).pack(side="right", padx=5)
        
        tk.Button(control_frame, text="☐ Deselect All", command=deselect_all,
                 bg="#FFD951", fg="black",
                 font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                 padx=12, pady=6).pack(side="right", padx=5)
        
        # Action buttons
        action_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        action_frame.pack(fill="x", padx=20, pady=15)
        
        def export_to_excel():
            """Export selected users to Excel"""
            import tempfile
            import subprocess
            import platform
            
            selected = []
            for item in tree.get_children():
                if tree.item(item, "text") == "☑":
                    values = tree.item(item, "values")
                    selected.append(values)
            
            if not selected:
                self.show_dark_messagebox("No Selection", "Please select at least one user to export.", "warning")
                return
            
            try:
                temp_dir = tempfile.gettempdir()
                filename = f"Inactive_Users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = os.path.join(temp_dir, filename)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inactive Users"
                
                # Header styling
                header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                headers = ["Name", "Username", "Job Title", "Last Login", "Days Inactive"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add data
                for row_num, (name, username, job_title, last_login) in enumerate(selected, 2):
                    ws.cell(row=row_num, column=1, value=name)
                    ws.cell(row=row_num, column=2, value=username)
                    ws.cell(row=row_num, column=3, value=job_title)
                    ws.cell(row=row_num, column=4, value=last_login)
                    
                    try:
                        last_login_date = datetime.strptime(last_login, DATE_FORMAT)
                        days_inactive = (datetime.now() - last_login_date).days
                        ws.cell(row=row_num, column=5, value=days_inactive)
                    except:
                        ws.cell(row=row_num, column=5, value="N/A")
                
                # Column widths
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 15
                
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'
                
                # Summary sheet
                summary_ws = wb.create_sheet(title="Summary")
                summary_ws['A1'] = "Inactive Users Report"
                summary_ws['A1'].font = Font(bold=True, size=14)
                summary_ws['A3'] = "Report Generated:"
                summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                summary_ws['A4'] = "Total Users:"
                summary_ws['B4'] = len(selected)
                summary_ws['A5'] = "Inactive Threshold:"
                summary_ws['B5'] = f"{INACTIVE_DAYS} days"
                
                wb.save(filepath)
                
                self.log_to_ui(f"\n✓ Excel report created: {filename}\n")
                
                # Open file
                if platform.system() == 'Windows':
                    os.startfile(filepath)
                elif platform.system() == 'Darwin':
                    subprocess.run(['open', filepath])
                else:
                    subprocess.run(['xdg-open', filepath])
                
                self.show_dark_messagebox("Export Complete", 
                                        f"Report opened in Excel!\n\nFile: {filename}",
                                        "info")
                
            except Exception as e:
                self.log_to_ui(f"\n✗ Export error: {str(e)}\n")
                self.show_dark_messagebox("Export Error", f"Failed to export:\n{str(e)}", "error")
        
        def start_deactivation():
            """Start deactivating selected users"""
            selected = []
            for item in tree.get_children():
                if tree.item(item, "text") == "☑":
                    values = tree.item(item, "values")
                    # Convert back to tuple format (name, username, jobtitle, lastlogin)
                    selected.append((values[0], values[1], values[2], values[3]))
            
            if not selected:
                self.show_dark_messagebox("No Selection", "Please select at least one user to deactivate.", "warning")
                return
            
            # Confirm
            if not self.show_dark_confirm("Confirm Deactivation", 
                                         f"Deactivate {len(selected)} selected user(s)?"):
                return
            
            dialog.destroy()
            
            # Run deactivation with the driver we already have
            self.run_scan_deactivation(selected, driver)
        
        tk.Button(action_frame, text="📊 Export to Excel", command=export_to_excel,
                 bg=self.bg_light, fg=self.text_light,
                 font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2",
                 width=18, height=2).pack(side="left", padx=(50, 10))
        
        tk.Button(action_frame, text="✓ Deactivate Selected", command=start_deactivation,
                 bg="#4CAF50", fg="white",
                 font=("Segoe UI", 12, "bold"), relief="flat", cursor="hand2",
                 width=22, height=2).pack(side="left", padx=10)
        
        def cancel_and_close():
            """Cancel and close browser"""
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            dialog.destroy()
            # Re-enable buttons
            self.update_progress(0, "Ready")
            self.start_scan_button.config(state="normal", text="▶ Start Scan")
            self.batch_deactivate_button.config(state="normal")
        
        tk.Button(action_frame, text="✕ Cancel", command=cancel_and_close,
                 bg="#ff4444", fg="white",
                 font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2",
                 width=12, height=2).pack(side="left", padx=10)
        
        dialog.wait_window()
    
    
    def show_spreadsheet_selection_dialog(self, users):
        """Show dialog to select users from spreadsheet and add tickets"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Users from Spreadsheet")
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        dialog.geometry("1000x700")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"1000x700+{x}+{y}")
        
        # Header
        tk.Label(dialog, text=f"{len(users)} Users from Spreadsheet", 
                font=("Segoe UI", 16, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.accent).pack(pady=15, padx=20)
        
        tk.Label(dialog, text="Select users and add ticket numbers for deactivation", 
                font=("Segoe UI", 10),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_dim).pack(pady=(0, 5), padx=20)
        
        # Instructions
        tk.Label(dialog, text="💡 Right-click Name or Username to exclude user from future scans", 
                font=("Segoe UI", 9, "italic"),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.accent).pack(pady=(0, 10), padx=20)
        
        # Treeview for user list with ticket column
        list_frame = tk.Frame(dialog, 
                             bg="#1a1a1a" if self.is_dark_mode else "#f5f5f5",
                             relief="solid", 
                             borderwidth=1)
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Style
        style = ttk.Style()
        style.theme_use('clam')  # Use clam theme for better dark mode support
        
        if self.is_dark_mode:
            # Dark mode styling
            style.configure("Spreadsheet.Treeview",
                           background="#2d2d2d",  # Dark grey background
                           foreground="#ffffff",  # White text
                           rowheight=32,
                           fieldbackground="#2d2d2d",  # Dark grey field background
                           borderwidth=0,
                           font=("Segoe UI", 9))
            style.configure("Spreadsheet.Treeview.Heading",
                           background="#1a1a1a",  # Darker header
                           foreground=self.accent,
                           font=("Segoe UI", 10, "bold"),
                           borderwidth=1,
                           relief="flat")
            style.map("Spreadsheet.Treeview",
                     background=[("selected", self.accent)],
                     foreground=[("selected", "white")])
            style.map("Spreadsheet.Treeview.Heading",
                     background=[("active", "#252525")])
        else:
            # Light mode styling
            style.configure("Spreadsheet.Treeview",
                           background="#ffffff",
                           foreground="#1a1a1a",
                           rowheight=32,
                           fieldbackground="#ffffff",
                           font=("Segoe UI", 9))
            style.configure("Spreadsheet.Treeview.Heading",
                           background="#e0e0e0",
                           foreground=self.accent,
                           font=("Segoe UI", 10, "bold"))
            style.map("Spreadsheet.Treeview",
                     background=[("selected", self.accent)],
                     foreground=[("selected", "white")])
        
        vsb = ttk.Scrollbar(list_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        
        tree = ttk.Treeview(list_frame,
                           columns=("name", "username", "ticket"),
                           show="tree headings",
                           style="Spreadsheet.Treeview",
                           yscrollcommand=vsb.set)
        
        vsb.config(command=tree.yview)
        
        tree.heading("#0", text="☑", anchor="center")
        tree.heading("name", text="Name", anchor="w")
        tree.heading("username", text="Username", anchor="w")
        tree.heading("ticket", text="Ticket / Reason", anchor="w")
        
        tree.column("#0", width=50, anchor="center")
        tree.column("name", width=200, anchor="w")
        tree.column("username", width=150, anchor="w")
        tree.column("ticket", width=400, anchor="w")
        
        tree.pack(side="left", fill="both", expand=True)
        
        # Store entry widgets for each row
        ticket_entries = {}
        
        # Add users with permanent entry boxes
        for user in users:
            item_id = tree.insert("", "end", 
                                 text="☑",
                                 values=(user['name'], user['username'], ""))
            
            # Create permanent entry widget for this row
            # We'll position it after the tree is visible
            ticket_entries[item_id] = None
        
        # Function to create and position entry widgets
        def create_ticket_entries():
            for item_id in ticket_entries.keys():
                try:
                    # Get position of ticket column for this item
                    bbox = tree.bbox(item_id, "ticket")
                    if not bbox:
                        continue
                    
                    x, y, width, height = bbox
                    
                    # Create entry widget
                    if self.is_dark_mode:
                        entry = tk.Entry(tree,
                                       font=("Segoe UI", 9),
                                       bg="#3d3d3d",
                                       fg="#ffffff",
                                       insertbackground="#ffffff",
                                       relief="solid",
                                       borderwidth=1)
                    else:
                        entry = tk.Entry(tree,
                                       font=("Segoe UI", 9),
                                       bg="#f8f8f8",
                                       fg="#1a1a1a",
                                       relief="solid",
                                       borderwidth=1)
                    
                    # Position entry over the ticket column
                    entry.place(x=x, y=y, width=width-2, height=height-2)
                    ticket_entries[item_id] = entry
                except:
                    pass
        
        # Create entries after dialog is visible
        dialog.update_idletasks()
        create_ticket_entries()
        
        # Update entries on scroll
        def on_scroll(*args):
            vsb.set(*args)
            # Reposition visible entries
            for item_id, entry in ticket_entries.items():
                if entry:
                    try:
                        bbox = tree.bbox(item_id, "ticket")
                        if bbox:
                            x, y, width, height = bbox
                            entry.place(x=x, y=y, width=width-2, height=height-2)
                        else:
                            entry.place_forget()
                    except:
                        pass
        
        tree.config(yscrollcommand=on_scroll)
        
        def toggle_user(event):
            region = tree.identify("region", event.x, event.y)
            column = tree.identify_column(event.x)
            
            # Only toggle checkbox on tree column
            if region == "tree" and column == "#0":
                item = tree.identify_row(event.y)
                if item:
                    current = tree.item(item, "text")
                    tree.item(item, text="☐" if current == "☑" else "☑")
        
        tree.bind("<Button-1>", toggle_user)
        
        # Right-click menu to exclude user from future scans
        def show_exclude_menu(event):
            # Get clicked item
            item = tree.identify_row(event.y)
            if not item:
                return
            
            # Get user name
            vals = tree.item(item, "values")
            name = vals[0]  # Name field
            
            # Create context menu
            menu = tk.Menu(tree, tearoff=0, bg=self.bg_card, fg=self.text_light)
            menu.add_command(label=f"Exclude '{name}' from future scans?", 
                           command=lambda: self.exclude_user_from_scans(name, dialog))
            menu.add_separator()
            menu.add_command(label="Cancel")
            
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
        
        tree.bind("<Button-3>", show_exclude_menu)  # Right-click
        
        # Bottom buttons
        button_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        button_frame.pack(fill="x", padx=20, pady=10)
        
        def proceed():
            selected = []
            for item in tree.get_children():
                if tree.item(item, "text") == "☑":
                    vals = tree.item(item, "values")
                    # Get ticket from entry widget
                    entry = ticket_entries.get(item)
                    ticket = entry.get().strip() if entry else ""
                    if not ticket:
                        ticket = "NO-TICKET"
                    selected.append((vals[1], ticket))  # (username, ticket)
            
            if not selected:
                self.show_dark_messagebox("No Users Selected", "Please select at least one user.", "warning")
                return
            
            # Confirm
            if self.show_dark_confirm("Confirm Deactivation",
                                     f"Deactivate {len(selected)} users?"):
                dialog.destroy()
                # Run deactivation
                self.run_spreadsheet_deactivation(selected)
        
        tk.Button(button_frame,
                 text=f"Deactivate Selected Users",
                 command=proceed,
                 bg="#4CAF50",
                 fg="white",
                 font=("Segoe UI", 11, "bold"),
                 relief="flat",
                 cursor="hand2",
                 padx=20,
                 pady=10).pack(side="right")
        
        tk.Button(button_frame,
                 text="Cancel",
                 command=dialog.destroy,
                 bg="#999999",
                 fg="white",
                 font=("Segoe UI", 11, "bold"),
                 relief="flat",
                 cursor="hand2",
                 padx=20,
                 pady=10).pack(side="right", padx=(0, 10))
    
    def run_spreadsheet_deactivation(self, selected_users):
        """Deactivate users from spreadsheet with their tickets"""
        
        # Show warning if not in headless mode
        if not self.headless_mode:
            self.show_dark_messagebox("Non-Headless Mode Active",
                                    "⚠️ Chrome is running in visible mode.\n\n"
                                    "Please DO NOT:\n"
                                    "• Minimize the Chrome window\n"
                                    "• Click away from Chrome\n"
                                    "• Close Chrome manually\n\n"
                                    "Keep Chrome visible during the deactivation process.",
                                    "warning")
        
        import threading
        
        def run():
            driver = None
            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.by import By
                import time
                
                self.log_to_ui("="*60)
                self.log_to_ui(f"DEACTIVATING {len(selected_users)} USERS FROM SPREADSHEET")
                self.log_to_ui("="*60 + "\n")
                
                # Create driver
                chrome_options = self.get_chrome_options()
                chrome_options.add_argument("--start-maximized")  # Full screen
                
                service = Service()
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.set_page_load_timeout(20)
                
                actions = ActionChains(driver)
                wait = WebDriverWait(driver, 10)
                
                # Login
                self.log_to_ui("→ Logging in...")
                driver.get(LOGIN_URL)
                
                self.automation_sleep(1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.username).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.password).perform()
                self.automation_sleep(0.2)
                
                login_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
                ))
                driver.execute_script("arguments[0].click();", login_button)
                self.automation_sleep(2)
                
                # Select program
                self.log_to_ui("→ Selecting program...")
                self.automation_sleep(1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys("0000").perform()
                self.automation_sleep(0.3)
                actions.send_keys(Keys.ARROW_DOWN).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.ENTER).perform()
                self.automation_sleep(0.2)
                
                submit_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
                ))
                driver.execute_script("arguments[0].click();", submit_button)
                self.automation_sleep(2)
                
                self.log_to_ui("  ✓ Logged in\n")
                
                # Process each user
                success_count = 0
                fail_count = 0
                
                for i, (username, ticket_reason) in enumerate(selected_users, 1):
                    try:
                        self.log_to_ui(f"[{i}/{len(selected_users)}] Processing {username}...")
                        
                        # Deactivate using same function as batch deactivate
                        # The function handles navigation, search, and deactivation
                        result = deactivate_user_by_ticket(driver, username, ticket_reason, self.automation_speed, self)
                        
                        if result:
                            self.log_to_ui(f"  ✓ Deactivated: {username} (Ticket: {ticket_reason})")
                            success_count += 1
                        else:
                            self.log_to_ui(f"  ✗ Failed: {username}")
                            fail_count += 1
                        
                        self.automation_sleep(0.5)  # Brief pause between users
                            
                    except Exception as e:
                        import traceback
                        self.log_to_ui(f"  ✗ Error deactivating {username}: {str(e)}")
                        self.log_to_ui(f"{traceback.format_exc()}")
                        fail_count += 1
                
                self.log_to_ui(f"\n{'='*60}")
                self.log_to_ui(f"BATCH COMPLETE")
                self.log_to_ui(f"  Total:   {len(selected_users)}")
                self.log_to_ui(f"  Success: {success_count}")
                self.log_to_ui(f"  Failed:  {fail_count}")
                self.log_to_ui(f"{'='*60}\n")
                
                # Close browser first so dialog is visible
                if driver:
                    try:
                        self.log_to_ui("→ Closing browser...")
                        driver.quit()
                        self.log_to_ui("  ✓ Browser closed\n")
                        driver = None  # Prevent double-close in finally
                    except:
                        pass
                
                # Show summary dialog
                self.show_dark_messagebox("Batch Complete", 
                                        f"Processed {len(selected_users)} users\n✓ Success: {success_count}\n✗ Failed: {fail_count}",
                                        "info")
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                self.log_to_ui(f"\n✗ ERROR: {str(e)}\n")
                self.log_to_ui(f"{error_details}\n")
                self.show_dark_messagebox("Error", str(e), "error")
            finally:
                if driver:
                    try:
                        self.log_to_ui("→ Closing browser...")
                        driver.quit()
                        self.log_to_ui("  ✓ Browser closed\n")
                    except:
                        pass
        
        threading.Thread(target=run, daemon=True).start()
    
    def _parse_date_safe(self, date_str):
        """Helper to parse date for sorting"""
        try:
            return datetime.strptime(date_str, DATE_FORMAT)
        except:
            return datetime.min
    
    def run_scan_deactivation(self, selected_users, driver):
        """Deactivate users from scan results"""
        import threading
        
        def run_deactivation():
            success_count = 0
            fail_count = 0
            failed_list = []
            
            try:
                self.log_to_ui("\n" + "="*60)
                self.log_to_ui(f"DEACTIVATING {len(selected_users)} USERS")
                self.log_to_ui("="*60 + "\n")
                
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.by import By
                import time
                
                actions = ActionChains(driver)
                wait = WebDriverWait(driver, 10)
                
                # Navigate to users page
                self.log_to_ui("→ Navigating to users page...")
                driver.get(get_url("/main/users", self))
                self.automation_sleep(1)
                driver.find_element(By.TAG_NAME, "body").click()
                self.automation_sleep(0.3)
                self.log_to_ui("  ✓ Ready\n")
                
                # Process each user
                for idx, (name, username, job_title, last_login) in enumerate(selected_users, 1):
                    progress = 10 + int((idx / len(selected_users)) * 85)
                    self.update_progress(progress, f"User {idx}/{len(selected_users)}")
                    
                    self.log_to_ui(f"[{idx}/{len(selected_users)}] {name} ({username})")
                    
                    # Use ticket from scan (you may want to modify this)
                    ticket = "SCAN-AUTO"  # Or prompt for ticket number
                    
                    success = deactivate_user_by_ticket(driver, username, ticket, self.automation_speed, self)
                    
                    if success:
                        success_count += 1
                        self.log_to_ui(f"  ✓ Deactivated successfully")
                    else:
                        fail_count += 1
                        failed_list.append((username, "Deactivation failed"))
                        self.log_to_ui(f"  ✗ Failed")
                    
                    self.log_to_ui("")
                
                self.update_progress(100, "Complete!")
                
                # Summary
                self.log_to_ui("="*60)
                self.log_to_ui("DEACTIVATION COMPLETE")
                self.log_to_ui(f"  Total users:     {len(selected_users)}")
                self.log_to_ui(f"  ✓ Successful:    {success_count}")
                self.log_to_ui(f"  ✗ Failed:        {fail_count}")
                self.log_to_ui("="*60 + "\n")
                
                # Show summary
                self.show_dark_messagebox("Deactivation Complete", 
                                        f"Processed {len(selected_users)} users\n✓ Success: {success_count}\n✗ Failed: {fail_count}",
                                        "info")
                
            except Exception as e:
                self.log_to_ui(f"\n✗ DEACTIVATION ERROR: {str(e)}\n")
                self.show_dark_messagebox("Error", f"Deactivation error: {str(e)}", "error")
            finally:
                # Close driver
                if driver:
                    try:
                        self.log_to_ui("→ Closing browser...")
                        driver.quit()
                        self.log_to_ui("  ✓ Browser closed\n")
                    except:
                        pass
                
                # Re-enable buttons
                self.update_progress(0, "Ready")
                self.start_scan_button.config(state="normal", text="▶ Start Scan")
                self.batch_deactivate_button.config(state="normal")
        
        threading.Thread(target=run_deactivation, daemon=True).start()
    
    def deactivate_with_updates(self, driver, username, ticket):
        """Deactivate user with UI progress updates"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        import time
        
        wait = WebDriverWait(driver, 10)
        actions = ActionChains(driver)
        
        try:
            self.update_progress(60, "Navigating to users page...")
            self.log_to_ui(f"→ Searching for user: {username}")
            
            # Navigate to users page
            driver.get(get_url("/main/users", self))
            self.automation_sleep(0.4)
            driver.find_element(By.TAG_NAME, "body").click()
            self.automation_sleep(0.05)
            
            # Search for user
            driver.execute_script("""
                var searchInput = document.querySelector('.table-search-input');
                if (searchInput) {
                    searchInput.focus();
                    searchInput.click();
                }
            """)
            self.automation_sleep(0.1)
            
            self.update_progress(65, "Searching...")
            
            # Type username and search
            actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
            self.automation_sleep(0.05)
            actions.send_keys(username).perform()
            self.automation_sleep(0.05)
            actions.send_keys(Keys.ENTER).perform()
            self.automation_sleep(1.0)
            
            self.update_progress(70, "User found, opening profile...")
            self.log_to_ui("  ✓ User found")
            
            # Find and click user link
            user_link = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "span.text-link")
            ))
            driver.execute_script("arguments[0].click();", user_link)
            self.automation_sleep(1.0)
            
            self.update_progress(75, "Executing deactivation sequence...")
            self.log_to_ui(f"→ Deactivating with Ticket {ticket}...")
            
            # Execute deactivation sequence
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.ARROW_DOWN).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.SPACE).perform()
            self.automation_sleep(0.1)
            
            self.update_progress(80, "Filling deactivation form...")
            
            actions.send_keys(f"Ticket {ticket} Deactivated").perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys("12@Deactivate").perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.SPACE).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.SPACE).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.SPACE).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.ENTER).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.ENTER).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys(Keys.TAB).perform()
            self.automation_sleep(0.1)
            actions.send_keys("12@Deactivate").perform()
            self.automation_sleep(0.1)
            
            self.update_progress(90, "Saving changes...")
            self.log_to_ui("  → Saving...")
            
            # Scroll and save
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.automation_sleep(0.1)
            
            save_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(text(), 'SAVE')]")
            ))
            driver.execute_script("arguments[0].click();", save_button)
            self.automation_sleep(0.8)
            
            self.log_to_ui("  ✓ Changes saved!")
            return True
            
        except Exception as e:
            self.log_to_ui(f"  ✗ Error: {str(e)}")
            print(f"✗ Deactivation failed: {str(e)}")
            return False
    
    def update_progress(self, percent, status_text):
        """Update progress bar and status text"""
        try:
            # Update status label with actual status text
            if hasattr(self, 'status_label'):
                self.status_label.config(text=status_text)
            
            # Update progress bar
            if hasattr(self, 'progress_bar'):
                self.progress_bar.place(relwidth=percent/100)
            
            # Update percent label
            if hasattr(self, 'progress_percent_label'):
                self.progress_percent_label.config(text=f"{percent}%")
            
            # Force UI update
            self.root.update_idletasks()
        except:
            pass
        
        # Print to console
        print(f"[{percent}%] {status_text}")
    
    def log_to_ui(self, message):
        """Log message to UI console and terminal"""
        print(message)
        
        # Also update UI log widget
        try:
            if hasattr(self, 'log_text'):
                self.log_text.config(state="normal")
                self.log_text.insert(tk.END, message + "\n")
                self.log_text.see(tk.END)
                self.log_text.config(state="disabled")
                self.root.update_idletasks()
        except:
            pass
    
    def show_dark_input(self, title, prompt):
        """Show input dialog with dark mode styling"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        # Center dialog
        dialog.geometry("400x150")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        tk.Label(dialog, text=prompt, 
                font=("Segoe UI", 11),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_light).pack(pady=20, padx=20)
        
        entry = tk.Entry(dialog, font=("Segoe UI", 12), width=30,
                        bg=self.bg_light if self.is_dark_mode else "#ffffff",
                        fg=self.text_light,
                        insertbackground=self.text_light if self.is_dark_mode else "#000000")
        entry.pack(pady=10, padx=20)
        entry.focus()
        
        result = {"value": None}
        
        def on_ok():
            result["value"] = entry.get()
            dialog.destroy()
        
        def on_cancel():
            dialog.destroy()
        
        entry.bind('<Return>', lambda e: on_ok())
        entry.bind('<Escape>', lambda e: on_cancel())
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="OK", command=on_ok, width=10,
                 bg=self.accent, fg="white" if not self.is_dark_mode else "black",
                 font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=on_cancel, width=10,
                 bg=self.bg_light, fg=self.text_light,
                 font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=5)
        
        dialog.wait_window()
        return result["value"]
    
    def show_dark_confirm(self, title, message):
        """Show yes/no dialog with dark mode styling"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        # Bigger dialog with more space
        dialog.geometry("450x200")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"450x200+{x}+{y}")
        
        # Content with more padding
        tk.Label(dialog, text=message, 
                font=("Segoe UI", 11),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_light,
                wraplength=400,
                justify="center").pack(pady=35, padx=25)
        
        result = {"value": False}
        
        def on_yes():
            result["value"] = True
            dialog.destroy()
        
        def on_no():
            dialog.destroy()
        
        # Buttons - normal size
        btn_frame = tk.Frame(dialog, bg=self.bg_card if self.is_dark_mode else "#ffffff")
        btn_frame.pack(pady=(0, 20))
        
        tk.Button(btn_frame, 
                 text="Yes", 
                 command=on_yes, 
                 width=10,
                 bg=self.accent, 
                 fg="white",
                 font=("Segoe UI", 10, "bold"), 
                 relief="flat", 
                 cursor="hand2",
                 padx=10,
                 pady=6).pack(side="left", padx=8)
        
        tk.Button(btn_frame, 
                 text="No", 
                 command=on_no, 
                 width=10,
                 bg="#999999", 
                 fg="white",
                 font=("Segoe UI", 10, "bold"), 
                 relief="flat", 
                 cursor="hand2",
                 padx=10,
                 pady=6).pack(side="left", padx=8)
        
        dialog.wait_window()
        return result["value"]
    
    def show_dark_messagebox(self, title, message, msg_type="info"):
        """Show message dialog with dark mode styling"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.configure(bg=self.bg_card if self.is_dark_mode else "#ffffff")
        dialog.transient(self.root)
        dialog.grab_set()
        
        if self.is_dark_mode:
            self.enable_dark_title_bar(dialog)
        
        # Calculate height based on message length
        estimated_lines = len(message) / 45  # More generous estimate
        min_height = 250  # Taller minimum
        base_height = max(min_height, int(estimated_lines * 28) + 180)  # More padding
        
        dialog.geometry(f"520x{base_height}")
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"520x{base_height}+{x}+{y}")
        
        # Icon based on type
        icons = {"info": "ℹ️", "error": "❌", "warning": "⚠️", "success": "✓"}
        icon = icons.get(msg_type, "ℹ️")
        
        # Content with more padding
        tk.Label(dialog, text=f"{icon} {message}", 
                font=("Segoe UI", 11),
                bg=self.bg_card if self.is_dark_mode else "#ffffff",
                fg=self.text_light,
                wraplength=450,
                justify="left").pack(pady=30, padx=30)
        
        def on_ok():
            dialog.destroy()
        
        # OK Button - normal size
        tk.Button(dialog, 
                 text="OK", 
                 command=on_ok, 
                 width=15,
                 bg=self.accent, 
                 fg="white",
                 font=("Segoe UI", 10, "bold"), 
                 relief="flat", 
                 cursor="hand2",
                 padx=15,
                 pady=8).pack(pady=(0, 20))
        
        dialog.bind('<Return>', lambda e: on_ok())
        dialog.bind('<Escape>', lambda e: on_ok())
        
        dialog.wait_window()
    
    def show_user_standards(self):
        self.clear_container()
        tool_frame = tk.Frame(self.container, bg=self.bg_dark)
        tool_frame.pack(fill="both", expand=True)
        
        # Header
        header_frame = tk.Frame(tool_frame, bg=self.bg_card, height=80)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button in top right
        tk.Button(header_frame, 
                  text="← Back to Menu", 
                  command=self.show_standards_scans_menu,
                  bg=self.bg_light, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(relx=1.0, x=-20, y=20, anchor="ne")
        
        tk.Label(header_frame, text="📋 User Standards Scan", font=("Segoe UI", 20, "bold"),
                bg=self.bg_card, fg=self.accent).pack(pady=20)
        
        # Main container
        main_container = tk.Frame(tool_frame, bg=self.bg_dark)
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Left side - Console log
        left_frame = tk.Frame(main_container, bg=self.bg_dark)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        tk.Label(left_frame, text="Console Log", font=("Segoe UI", 12, "bold"),
                bg=self.bg_dark, fg=self.text_light).pack(anchor="w", pady=(0, 5))
        
        log_frame = tk.Frame(left_frame, bg=self.bg_card)
        log_frame.pack(fill="both", expand=True)
        
        self.standards_log_text = scrolledtext.ScrolledText(log_frame, bg=self.bg_card, fg=self.text_light,
                          font=("Consolas", 10), wrap="word", relief="flat",
                          padx=10, pady=10, state="normal")
        self.standards_log_text.pack(fill="both", expand=True)
        
        # Right side - Stats and controls
        right_frame = tk.Frame(main_container, bg=self.bg_dark, width=300)
        right_frame.pack(side="right", fill="y")
        right_frame.pack_propagate(False)
        
        stats_content = tk.Frame(right_frame, bg=self.bg_card)
        stats_content.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Header with Stop link
        header_frame = tk.Frame(stats_content, bg=self.bg_card)
        header_frame.pack(fill="x", pady=(15, 10))
        
        tk.Label(header_frame, text="Scan Statistics", font=("Segoe UI", 12, "bold"),
                bg=self.bg_card, fg=self.accent).pack(side="left", padx=(15, 0))
        
        # Stop scan link on the right
        self.standards_stop_link = tk.Label(header_frame, text="⏹ Stop", 
                                           font=("Segoe UI", 9, "bold"),
                                           bg=self.bg_card,
                                           fg="#ff4444", cursor="hand2", padx=10)
        self.standards_stop_link.pack(side="right")
        self.standards_stop_link.bind("<Button-1>", lambda e: self.stop_standards_scan())
        self.standards_stop_link.pack_forget()  # Hidden by default
        
        # Stats - Store references
        stats_frame = tk.Frame(stats_content, bg=self.bg_card)
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        # Pages Scanned
        stat_frame1 = tk.Frame(stats_frame, bg=self.bg_card)
        stat_frame1.pack(fill="x", pady=5)
        tk.Label(stat_frame1, text="Pages Scanned:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(side="left")
        self.standards_pages_label = tk.Label(stat_frame1, text="0", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg=self.text_light, anchor="e")
        self.standards_pages_label.pack(side="right")
        
        # Total Users
        stat_frame2 = tk.Frame(stats_frame, bg=self.bg_card)
        stat_frame2.pack(fill="x", pady=5)
        tk.Label(stat_frame2, text="Total Users:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(side="left")
        self.standards_total_label = tk.Label(stat_frame2, text="0", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg=self.text_light, anchor="e")
        self.standards_total_label.pack(side="right")
        
        # Compliant
        stat_frame3 = tk.Frame(stats_frame, bg=self.bg_card)
        stat_frame3.pack(fill="x", pady=5)
        tk.Label(stat_frame3, text="✓ Compliant:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(side="left")
        self.standards_compliant_label = tk.Label(stat_frame3, text="0", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg=self.text_light, anchor="e")
        self.standards_compliant_label.pack(side="right")
        
        # Non-Compliant
        stat_frame4 = tk.Frame(stats_frame, bg=self.bg_card)
        stat_frame4.pack(fill="x", pady=5)
        tk.Label(stat_frame4, text="✗ Non-Compliant:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(side="left")
        self.standards_noncompliant_label = tk.Label(stat_frame4, text="0", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg=self.text_light, anchor="e")
        self.standards_noncompliant_label.pack(side="right")
        
        # Separator
        tk.Frame(stats_content, bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                height=2).pack(fill="x", padx=20, pady=15)
        
        # Status
        tk.Label(stats_content, text="Status:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", padx=20)
        self.standards_status_label = tk.Label(stats_content, text="Ready", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg="#4CAF50", anchor="w")
        self.standards_status_label.pack(fill="x", padx=20, pady=5)
        
        # Progress
        tk.Label(stats_content, text="Progress:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", padx=20, pady=(15, 2))
        
        progress_frame = tk.Frame(stats_content,
                                 bg=self.bg_light if self.is_dark_mode else "#e0e0e0",
                                 height=25)
        progress_frame.pack(fill="x", padx=20, pady=(0, 5))
        progress_frame.pack_propagate(False)
        
        self.standards_progress_bar = tk.Frame(progress_frame, bg=self.accent, height=25)
        self.standards_progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        self.standards_progress_label = tk.Label(stats_content, text="0%", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="center")
        self.standards_progress_label.pack(fill="x", padx=20)
        
        # Buttons
        button_color = "#0052CC" if self.is_dark_mode else "#0066ff"
        
        self.standards_start_button = tk.Button(stats_content, text="▶ Start Scan",
                 command=self.start_user_standards_scan,
                 bg=button_color, fg="black", font=("Segoe UI", 11, "bold"),
                 relief="flat", cursor="hand2", pady=10)
        self.standards_start_button.pack(fill="x", padx=20, pady=(20, 0))
        
        tk.Button(stats_content, text="📊 Export Report", state="disabled",
                 bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                 fg="black", font=("Segoe UI", 10, "bold"),
                 relief="flat", pady=10).pack(fill="x", padx=20, pady=5)
        
        # Stop flag
        self.should_stop_standards_scan = False
        
        self.current_screen = tool_frame
        self.root.title("SMART Assistant - User Standards Scan")
    
    def stop_standards_scan(self):
        """Stop the standards scan"""
        self.should_stop_standards_scan = True
        # Grey out the stop text
        self.standards_stop_link.config(text="Stopping...", fg="#999999")
    
    def standards_log(self, message):
        """Log message to standards scan console"""
        if hasattr(self, 'standards_log_text'):
            self.standards_log_text.insert(tk.END, message + "\n")
            self.standards_log_text.see(tk.END)
            self.root.update_idletasks()
    
    def start_user_standards_scan(self):
        """Start the user standards scan - checks username and job title compliance"""
        import threading
        
        logger.info("="*60)
        logger.info("STANDARDS SCAN STARTED - Username compliance check")
        logger.info(f"Settings: Headless={self.headless_mode}, Speed={self.automation_speed}, Environment={self.environment}")
        logger.info("="*60)
        
        # Disable button
        self.standards_start_button.config(state="disabled", text="Scanning...")
        self.standards_status_label.config(text="Scanning...", fg="#FFD951")
        
        # Show stop scan button
        self.should_stop_standards_scan = False
        self.standards_stop_link.pack(side="right")
        self.standards_stop_link.config(text="⏹ Stop", fg="#ff4444")
        
        # Show warning if not in headless mode
        if not self.headless_mode:
            self.show_dark_messagebox("Non-Headless Mode Active",
                                    "⚠️ Chrome is running in visible mode.\n\n"
                                    "Please DO NOT:\n"
                                    "• Minimize the Chrome window\n"
                                    "• Click away from Chrome\n"
                                    "• Close Chrome manually\n\n"
                                    "Keep Chrome visible during the standards scan.",
                                    "warning")
        
        def check_username_compliance(username):
            """Check if username follows standard format: xxxxflast"""
            import re
            
            # Pattern: 4 digits, then 1 letter (first initial), then letters (last name)
            pattern = r'^\d{4}[a-zA-Z][a-zA-Z]+$'
            
            if not username or len(username) < 6:
                return False, "Too short (min 6 chars)"
            
            if not re.match(pattern, username):
                if not username[:4].isdigit():
                    return False, "Missing 4-digit location code"
                else:
                    return False, "Invalid format (should be: xxxxflast)"
            
            return True, "Compliant"
        
        def run_scan():
            driver = None
            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                
                self.standards_log("="*60)
                self.standards_log("STARTING USERNAME STANDARDS SCAN")
                self.standards_log("="*60)
                self.standards_log("")
                self.standards_log("Standard Format: xxxxflast")
                self.standards_log("  xxxx = 4-digit location code")
                self.standards_log("  f = first initial")
                self.standards_log("  last = last name")
                self.standards_log("")
                
                # Launch browser (SAME AS DEACTIVATION SCAN)
                self.standards_log("→ Launching Chrome...")
                chrome_options = self.get_chrome_options()
                chrome_options.add_argument("--start-maximized")
                
                service = Service()
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.set_page_load_timeout(20)
                
                wait = WebDriverWait(driver, 10)
                actions = ActionChains(driver)
                
                # Login
                self.standards_log("→ Logging in...")
                self.standards_progress_bar.place(relwidth=0.1)
                self.standards_progress_label.config(text="10%")
                
                driver.get(LOGIN_URL)
                self.automation_sleep(1)
                
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.username).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys(self.password).perform()
                self.automation_sleep(0.2)
                
                login_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
                ))
                driver.execute_script("arguments[0].click();", login_button)
                self.automation_sleep(2)
                
                # Select program
                self.standards_log("→ Selecting program...")
                self.automation_sleep(1)
                actions.send_keys(Keys.TAB).perform()
                self.automation_sleep(0.1)
                actions.send_keys("0000").perform()
                self.automation_sleep(0.3)
                actions.send_keys(Keys.ARROW_DOWN).perform()
                self.automation_sleep(0.1)
                actions.send_keys(Keys.ENTER).perform()
                self.automation_sleep(0.2)
                
                submit_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
                ))
                driver.execute_script("arguments[0].click();", submit_button)
                self.automation_sleep(2)
                
                self.standards_log("  ✓ Logged in")
                self.standards_progress_bar.place(relwidth=0.2)
                self.standards_progress_label.config(text="20%")
                
                # Navigate to users page
                self.standards_log("→ Navigating to users page...")
                driver.get(get_url("/main/users", self))
                self.automation_sleep(1)
                driver.find_element(By.TAG_NAME, "body").click()
                self.automation_sleep(0.3)
                
                self.standards_log("  ✓ On users page")
                
                # Sort by Last Login (SAME AS DEACTIVATION SCAN)
                self.standards_log("→ Sorting by Last Login (descending)...")
                self.standards_progress_bar.place(relwidth=0.25)
                self.standards_progress_label.config(text="25%")
                
                try:
                    # Scroll to top first
                    driver.execute_script("window.scrollTo(0, 0);")
                    self.automation_sleep(0.5)
                    
                    # Try clicking the sort arrow
                    result = driver.execute_script("""
                        var path = document.querySelector("#root > div.collapsed.AdminLayout_admin_layout_container__8z6D1 > div.AdminLayout_admin_body_container__qHWPk > div > div.content-container > div > div.maintable-content.table-responsive > div.table-scroll > table > thead > tr > th:nth-child(7) > div > svg > path");
                        if (path) {
                            var event = new MouseEvent('click', {
                                view: window,
                                bubbles: true,
                                cancelable: true
                            });
                            path.dispatchEvent(event);
                            return true;
                        }
                        return false;
                    """)
                    
                    if result:
                        self.automation_sleep(1.0)
                        self.standards_log("  ✓ Sorted by Last Login (descending)")
                    else:
                        # Fallback
                        headers = driver.find_elements(By.TAG_NAME, "th")
                        for header in headers:
                            if "last login" in header.text.lower():
                                driver.execute_script("arguments[0].click();", header)
                                self.automation_sleep(0.5)
                                driver.execute_script("arguments[0].click();", header)
                                self.automation_sleep(0.5)
                                self.standards_log("  ✓ Sorted by Last Login (descending)")
                                break
                        else:
                            self.standards_log("  ! Could not sort - continuing anyway")
                except Exception as e:
                    self.standards_log(f"  ! Sorting failed: {str(e)} - continuing anyway")
                
                self.standards_log("")
                
                # Set page size to 100 (EXACT SAME AS DEACTIVATION SCAN)
                self.standards_log("→ Setting page size to 100 users per page...")
                self.standards_progress_bar.place(relwidth=0.3)
                self.standards_progress_label.config(text="30%")
                
                try:
                    self.automation_sleep(0.5)
                    
                    # Shift+Tab 14 times to reach page size dropdown
                    self.standards_log("  → Navigating to page size dropdown (Shift+Tab 14x)...")
                    for i in range(14):
                        actions.key_down(Keys.SHIFT).send_keys(Keys.TAB).key_up(Keys.SHIFT).perform()
                        self.automation_sleep(0.15)  # Increased from 0.05
                    
                    self.automation_sleep(0.3)  # Increased from 0.2
                    
                    # Space to open dropdown
                    self.standards_log("  → Opening dropdown (Space)...")
                    actions.send_keys(Keys.SPACE).perform()
                    self.automation_sleep(0.3)  # Increased from 0.2
                    
                    # Down arrow twice to select 100
                    self.standards_log("  → Selecting 100 (Down Down)...")
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.3)  # Increased from 0.1
                    actions.send_keys(Keys.ARROW_DOWN).perform()
                    self.automation_sleep(0.3)  # Increased from 0.1
                    
                    # Enter to confirm
                    self.standards_log("  → Confirming selection (Enter)...")
                    actions.send_keys(Keys.ENTER).perform()
                    self.automation_sleep(1.5)
                    
                    self.standards_log("  ✓ Page size set to 100")
                    
                except Exception as e:
                    self.standards_log(f"  ! Page size change failed: {str(e)}")
                    self.standards_log("  ! Continuing with default page size")
                
                self.standards_log("")
                self.standards_log("→ Starting scan...")
                self.standards_log("")
                
                # Scan variables
                page = 1
                total_users = 0
                compliant = 0
                non_compliant = 0
                previous_page_hash = None
                
                # Scan pages (up to 50)
                while page <= 50:
                    # Check if user clicked stop
                    if self.should_stop_standards_scan:
                        self.standards_log("")
                        self.standards_log("="*60)
                        self.standards_log("SCAN STOPPED BY USER")
                        self.standards_log("="*60)
                        break
                    
                    try:
                        # Check if Chrome is still open
                        try:
                            driver.current_url
                        except:
                            self.standards_log("")
                            self.standards_log("="*60)
                            self.standards_log("SCAN STOPPED - Browser closed")
                            self.standards_log("="*60)
                            break
                        
                        # Wait for table
                        table = wait.until(EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "table")
                        ))
                        rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
                        
                        if len(rows) == 0:
                            self.standards_log(f"Page {page}: No rows found - stopping")
                            break
                        
                        # Duplicate check
                        if len(rows) >= 3:
                            page_hash = hash(tuple(row.text for row in rows[:3]))
                            if previous_page_hash == page_hash and page > 1:
                                self.standards_log(f"Page {page}: Duplicate detected - stopping")
                                break
                            previous_page_hash = page_hash
                        
                        page_non_compliant = 0
                        
                        # Scan each row
                        for row in rows:
                            try:
                                cells = row.find_elements(By.TAG_NAME, "td")
                                if len(cells) < 4:
                                    continue
                                
                                # Extract data (Name, Username, Job Title, Last Login)
                                name = cells[1].text.strip()
                                username = cells[2].text.strip()
                                job_title = cells[3].text.strip()
                                
                                # Skip deactivated users
                                name_lower = name.lower()
                                if any(x in name_lower for x in ["deactivat", "deact", "terminat", "termin", "(term", "(deact", "end access"]):
                                    continue
                                
                                # Skip clients
                                if "client" in job_title.lower():
                                    continue
                                
                                total_users += 1
                                
                                # Check username compliance
                                is_compliant, reason = check_username_compliance(username)
                                
                                # Check job title issues
                                job_lower = job_title.lower().strip()
                                
                                # Check for N/A or empty
                                if not job_title or job_lower in ['n/a', 'na', 'none', '-', 'null']:
                                    is_compliant = False
                                    if reason == "Compliant":
                                        reason = "No job title (N/A)"
                                    else:
                                        reason = f"{reason}; No job title"
                                
                                # Check for multiple jobs
                                elif 'custom title' not in job_lower:
                                    if ',' in job_title or ' and ' in job_lower:
                                        is_compliant = False
                                        if reason == "Compliant":
                                            reason = "Multiple job titles"
                                        else:
                                            reason = f"{reason}; Multiple jobs"
                                
                                if is_compliant:
                                    compliant += 1
                                else:
                                    non_compliant += 1
                                    page_non_compliant += 1
                            
                            except:
                                pass
                        
                        self.standards_log(f"Page {page}: {page_non_compliant} non-compliant found")
                        
                        # Update stats
                        self.standards_pages_label.config(text=str(page))
                        self.standards_total_label.config(text=str(total_users))
                        self.standards_compliant_label.config(text=str(compliant))
                        self.standards_noncompliant_label.config(text=str(non_compliant))
                        
                        # Update progress
                        progress = 30 + int((page / 50) * 70)
                        self.standards_progress_bar.place(relwidth=progress/100)
                        self.standards_progress_label.config(text=f"{progress}%")
                        self.root.update_idletasks()
                        
                        # Navigate to next page (SAME AS DEACTIVATION SCAN)
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        self.automation_sleep(0.3)  # Increased from 0.1
                        actions.send_keys(Keys.TAB * 2).perform()
                        self.automation_sleep(0.4)  # Increased from 0.15
                        actions.send_keys(Keys.ENTER).perform()
                        self.automation_sleep(0.3)  # Increased from 0.1
                        actions.send_keys(Keys.ARROW_DOWN).perform()
                        self.automation_sleep(0.3)  # Increased from 0.1
                        actions.send_keys(Keys.ENTER).perform()
                        self.automation_sleep(1.0)  # Increased from 0.5
                        
                        page += 1
                    
                    except Exception as e:
                        self.standards_log(f"Error on page {page}: {str(e)}")
                        break
                
                # Complete
                self.standards_progress_bar.place(relwidth=1.0)
                self.standards_progress_label.config(text="100%")
                
                self.standards_log("")
                self.standards_log("="*60)
                self.standards_log("SCAN COMPLETE")
                self.standards_log("="*60)
                self.standards_log(f"  Pages scanned:     {page - 1}")
                self.standards_log(f"  Total users:       {total_users}")
                self.standards_log(f"  ✓ Compliant:       {compliant}")
                self.standards_log(f"  ✗ Non-compliant:   {non_compliant}")
                self.standards_log("="*60)
                
                self.standards_status_label.config(text="Complete", fg="#4CAF50")
                
                # Close browser
                if driver:
                    driver.quit()
                    self.standards_log("\n→ Browser closed")
                
                self.show_dark_messagebox("Scan Complete",
                                        f"Scanned {total_users} users across {page - 1} pages.\n\n"
                                        f"✓ Compliant: {compliant}\n"
                                        f"✗ Non-Compliant: {non_compliant}",
                                        "success")
                
            except Exception as e:
                import traceback
                self.standards_log(f"\n✗ ERROR: {str(e)}")
                self.standards_log(traceback.format_exc())
                self.standards_status_label.config(text="Error", fg="#ff4444")
                
                if driver:
                    try:
                        driver.quit()
                    except:
                        pass
                
                self.show_dark_messagebox("Scan Error", f"Error during scan:\n{str(e)}", "error")
            
            finally:
                # Re-enable button and hide stop scan
                self.standards_start_button.config(state="normal", text="▶ Start Scan")
                self.standards_stop_link.pack_forget()
        
        threading.Thread(target=run_scan, daemon=True).start()
    
    def show_equipment_standards(self):
        """Equipment Standards Scan page - mirrors User Standards UI"""
        self.clear_container()
        tool_frame = tk.Frame(self.container, bg=self.bg_dark)
        tool_frame.pack(fill="both", expand=True)
        
        # Header
        header_frame = tk.Frame(tool_frame, bg=self.bg_card, height=80)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button in top right
        tk.Button(header_frame, 
                  text="← Back to Menu", 
                  command=self.show_standards_scans_menu,
                  bg=self.bg_light, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(relx=1.0, x=-20, y=20, anchor="ne")
        
        tk.Label(header_frame, text="🔧 Equipment Standards Scan", font=("Segoe UI", 20, "bold"),
                bg=self.bg_card, fg=self.accent).pack(pady=20)
        
        # Main container
        main_container = tk.Frame(tool_frame, bg=self.bg_dark)
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Left side - Console log
        left_frame = tk.Frame(main_container, bg=self.bg_dark)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        tk.Label(left_frame, text="Console Log", font=("Segoe UI", 12, "bold"),
                bg=self.bg_dark, fg=self.text_light).pack(anchor="w", pady=(0, 5))
        
        log_frame = tk.Frame(left_frame, bg=self.bg_card)
        log_frame.pack(fill="both", expand=True)
        
        log_text = tk.Text(log_frame, bg=self.bg_card, fg=self.text_light,
                          font=("Consolas", 10), wrap="word", relief="flat",
                          padx=10, pady=10, state="disabled")
        log_text.pack(fill="both", expand=True)
        
        # Right side - Stats and controls
        right_frame = tk.Frame(main_container, bg=self.bg_dark, width=300)
        right_frame.pack(side="right", fill="y")
        right_frame.pack_propagate(False)
        
        stats_content = tk.Frame(right_frame, bg=self.bg_card)
        stats_content.pack(fill="both", expand=True, padx=5, pady=5)
        
        tk.Label(stats_content, text="Scan Statistics", font=("Segoe UI", 12, "bold"),
                bg=self.bg_card, fg=self.accent).pack(pady=(15, 10))
        
        # Stats
        stats_frame = tk.Frame(stats_content, bg=self.bg_card)
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        for label in ["Pages Scanned:", "Total Equipment:", "✓ Compliant:", "✗ Non-Compliant:"]:
            stat_frame = tk.Frame(stats_frame, bg=self.bg_card)
            stat_frame.pack(fill="x", pady=5)
            tk.Label(stat_frame, text=label, font=("Segoe UI", 9),
                    bg=self.bg_card, fg=self.text_dim, anchor="w").pack(side="left")
            tk.Label(stat_frame, text="0", font=("Segoe UI", 9, "bold"),
                    bg=self.bg_card, fg=self.text_light, anchor="e").pack(side="right")
        
        # Separator
        tk.Frame(stats_content, bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                height=2).pack(fill="x", padx=20, pady=15)
        
        # Status
        tk.Label(stats_content, text="Status:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", padx=20)
        tk.Label(stats_content, text="Coming Soon", font=("Segoe UI", 9, "bold"),
                bg=self.bg_card, fg="#FFD951", anchor="w").pack(fill="x", padx=20, pady=5)
        
        # Progress
        tk.Label(stats_content, text="Progress:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", padx=20, pady=(15, 2))
        
        progress_frame = tk.Frame(stats_content,
                                 bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                                 height=25)
        progress_frame.pack(fill="x", padx=20, pady=(0, 5))
        progress_frame.pack_propagate(False)
        
        progress_bar = tk.Frame(progress_frame, bg=self.accent, height=25)
        progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        tk.Label(stats_content, text="0%", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="center").pack(fill="x", padx=20)
        
        # Buttons (disabled) - Darker in dark mode with light grey text
        button_color_dark = "#0052CC" if self.is_dark_mode else "#0066ff"
        button_text_color = "#e0e0e0" if self.is_dark_mode else "black"
        
        tk.Button(stats_content, text="▶ Start Scan", state="disabled",
                 bg=button_color_dark, fg=button_text_color, font=("Segoe UI", 11, "bold"),
                 relief="flat", pady=10).pack(fill="x", padx=20, pady=(20, 0))
        
        tk.Button(stats_content, text="📊 Export Report", state="disabled",
                 bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                 fg="black", font=("Segoe UI", 10, "bold"),
                 relief="flat", pady=10).pack(fill="x", padx=20, pady=5)
        
        self.current_screen = tool_frame
        self.root.title("SMART Assistant - Equipment Standards Scan")
    
    def show_bbb_add_user(self):
        self.clear_container()
        self.current_screen = BBBAddUserScreen(self.container, self)
        self.root.title("SMART Assistant - BBB Add User")
        
        tk.Label(stats_panel, text="Setup Progress", font=("Segoe UI", 11, "bold"),
                bg=self.bg_card if self.is_dark_mode else "#fafaf8",
                fg=self.accent, anchor="w", padx=10, pady=8).pack(fill="x")
        
        stats_content = tk.Frame(stats_panel, bg=self.bg_card)
        stats_content.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Status
        tk.Label(stats_content, text="Status:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        tk.Label(stats_content, text="Coming Soon", font=("Segoe UI", 12, "bold"),
                bg=self.bg_card, fg="#FFD951", anchor="w").pack(fill="x", pady=(0, 15))
        
        # Progress
        tk.Label(stats_content, text="Progress:", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="w").pack(fill="x", pady=(0, 2))
        
        progress_frame = tk.Frame(stats_content, 
                                 bg=self.bg_card if self.is_dark_mode else "#fafaf8", 
                                 height=25)
        progress_frame.pack(fill="x", pady=(0, 5))
        progress_frame.pack_propagate(False)
        
        progress_bar = tk.Frame(progress_frame, bg=self.accent, height=25)
        progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        tk.Label(stats_content, text="0%", font=("Segoe UI", 9),
                bg=self.bg_card, fg=self.text_dim, anchor="center").pack(fill="x")
        
        # Action buttons (disabled)
        tk.Button(stats_content, text="📊 Add From Spreadsheet",
                 state="disabled", bg=("#0052CC" if self.is_dark_mode else "#0066ff"), fg=("#e0e0e0" if self.is_dark_mode else "black"),
                 font=("Segoe UI", 11, "bold"), relief="flat",
                 pady=10, highlightthickness=2,
                 highlightbackground="#333333").pack(fill="x", pady=(20, 5))
        
        self.current_screen = tool_frame
        self.root.title("SMART Assistant - BBB Add User")
    
    def show_sms_add_user(self):
        self.clear_container()
        self.current_screen = SMSAddUserScreen(self.container, self)
        self.root.title("SMART Assistant - SMS Add User")
    
    def on_closing(self):
        """Handle window close"""
        if self.show_dark_confirm("Quit", "Quit SMART Assistant?"):
            # Cleanup driver
            if hasattr(self, 'driver') and self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
            
            # Cleanup screen
            if self.current_screen and hasattr(self.current_screen, 'destroy'):
                try:
                    self.current_screen.destroy()
                except:
                    pass
            self.root.destroy()
    
    
    def run(self):
        """Start the application"""
        self.root.mainloop()


    def destroy(self):
        self.frame.destroy()




# ======================================================================
# LOGIN SCREEN
# ======================================================================

class LoginScreen:
    """Modern login screen with progress bar"""
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.build_ui()
    
    def build_ui(self):
        # Center container
        center = tk.Frame(self.frame, bg=self.app.bg_dark)
        center.place(relx=0.5, rely=0.5, anchor="center")
        
        # Logo/Title
        tk.Label(center, text="SMART ASSISTANT", 
                font=("Segoe UI", 42, "bold"),
                bg=self.app.bg_dark, fg=self.app.accent).pack(pady=(0, 15))
        
        tk.Label(center, text="User Authentication", 
                font=("Segoe UI", 16),
                bg=self.app.bg_dark, fg=self.app.text_dim).pack(pady=(0, 50))
        
        # Login card
        card = tk.Frame(center, bg=self.app.bg_card, relief="solid", borderwidth=1)
        card.pack(padx=60, pady=30)
        
        card_content = tk.Frame(card, bg=self.app.bg_card)
        card_content.pack(padx=50, pady=50)
        
        # Username
        # Username
        tk.Label(card_content, text="Username", font=("Segoe UI", 12, "bold"),
                bg=self.app.bg_card, fg=self.app.text_light, anchor="w").pack(fill="x", pady=(0, 8))
        
        username_frame = tk.Frame(card_content, bg="#ffffff" if not self.app.is_dark_mode else "#2a2a2a",
                                  relief="solid", borderwidth=1)
        username_frame.pack(fill="x", pady=(0, 25))
        
        self.username_entry = tk.Entry(username_frame, font=("Segoe UI", 12), width=35,
                                        bg="#ffffff" if not self.app.is_dark_mode else "#2a2a2a",
                                        fg=self.app.text_light, relief="flat", borderwidth=0)
        self.username_entry.pack(side="left", fill="both", expand=True, padx=15, pady=12)
        self.username_entry.focus()
        
        # Password
        tk.Label(card_content, text="Password", font=("Segoe UI", 12, "bold"),
                bg=self.app.bg_card, fg=self.app.text_light, anchor="w").pack(fill="x", pady=(0, 8))
        
        password_container = tk.Frame(card_content, bg=self.app.bg_card)
        password_container.pack(fill="x", pady=(0, 30))
        
        password_frame = tk.Frame(password_container, bg="#ffffff" if not self.app.is_dark_mode else "#2a2a2a",
                                  relief="solid", borderwidth=1)
        password_frame.pack(side="left", fill="both", expand=True)
        
        self.password_entry = tk.Entry(password_frame, font=("Segoe UI", 12), width=35, show="●",
                                        bg="#ffffff" if not self.app.is_dark_mode else "#2a2a2a",
                                        fg=self.app.text_light, relief="flat", borderwidth=0)
        self.password_entry.pack(side="left", fill="both", expand=True, padx=15, pady=12)
        
        # Eye toggle button
        self.password_visible = False
        self.eye_button = tk.Button(password_frame, text="👁", font=("Segoe UI", 12),
                                    bg="#ffffff" if not self.app.is_dark_mode else "#2a2a2a",
                                    fg=self.app.text_dim, relief="flat", borderwidth=0,
                                    cursor="hand2", command=self.toggle_password_visibility)
        self.eye_button.pack(side="right", padx=10)
        
        # Bind Enter key
        self.username_entry.bind("<Return>", lambda e: self.password_entry.focus())
        self.password_entry.bind("<Return>", lambda e: self.verify_login())
        
        # Status bar
        self.status_frame = tk.Frame(card_content, bg=self.app.bg_card, height=30)
        self.status_frame.pack(fill="x", pady=(0, 15))
        self.status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(self.status_frame, text="", font=("Segoe UI", 10),
                                     bg=self.app.bg_card, fg="#ff4444")
        self.status_label.pack()
        
        # Progress bar (hidden initially)
        self.progress_frame = tk.Frame(card_content, bg="#e0e0e0", height=4)
        self.progress_bar = tk.Frame(self.progress_frame, bg=self.app.accent, height=4)
        
        # Remember Me checkbox
        remember_frame = tk.Frame(card_content, bg=self.app.bg_card)
        remember_frame.pack(fill="x", pady=(0, 20))
        
        self.remember_var = tk.BooleanVar(value=False)
        self.remember_check = tk.Checkbutton(remember_frame,
                                             text="Remember Me",
                                             variable=self.remember_var,
                                             bg=self.app.bg_card,
                                             fg=self.app.text_light,
                                             selectcolor=self.app.bg_dark,
                                             activebackground=self.app.bg_card,
                                             activeforeground=self.app.text_light,
                                             font=("Segoe UI", 10),
                                             cursor="hand2")
        self.remember_check.pack(anchor="w")
        
        # Load saved credentials if they exist
        self.load_saved_credentials()
        
        # Login button
        button_container = tk.Frame(card_content, bg=self.app.bg_card)
        button_container.pack(fill="x", padx=60, pady=(0, 0))  # Add padding to make button narrower
        
        # Login button - Darker in dark mode to match tools
        login_button_color = "#0052CC" if self.app.is_dark_mode else "#0066ff"  # 20% darker
        
        self.login_button = tk.Button(button_container, text="Login →", command=self.verify_login,
                 bg=login_button_color, fg="white", font=("Segoe UI", 11, "bold"),
                 relief="flat", cursor="hand2", borderwidth=0,
                 activebackground=login_button_color, activeforeground="white")
        self.login_button.pack(fill="x", ipady=8)  # ipady 12→8 (40% smaller)
    
    
    def toggle_password_visibility(self):
        """Toggle password visibility"""
        self.password_visible = not self.password_visible
        if self.password_visible:
            self.password_entry.config(show="")
            self.eye_button.config(text="👁‍🗨")  # Eye with line through it
        else:
            self.password_entry.config(show="●")
            self.eye_button.config(text="👁")
    
    def load_saved_credentials(self):
        """Load saved credentials from settings if Remember Me was checked"""
        try:
            settings_path = os.path.join(os.path.expanduser("~"), "smart_assistant_settings.json")
            if os.path.exists(settings_path):
                with open(settings_path, 'r') as f:
                    settings = json.load(f)
                    saved_username = settings.get("saved_username", "")
                    saved_password = settings.get("saved_password", "")
                    
                    if saved_username and saved_password:
                        self.username_entry.insert(0, saved_username)
                        self.password_entry.insert(0, saved_password)
                        self.remember_var.set(True)
                        logger.info("Loaded saved credentials")
        except Exception as e:
            logger.error(f"Error loading saved credentials: {e}")
    
    def save_credentials(self, username, password):
        """Save credentials to settings file if Remember Me is checked"""
        try:
            settings_path = os.path.join(os.path.expanduser("~"), "smart_assistant_settings.json")
            
            # Load existing settings
            settings = {}
            if os.path.exists(settings_path):
                with open(settings_path, 'r') as f:
                    settings = json.load(f)
            
            # Update credentials
            if self.remember_var.get():
                settings["saved_username"] = username
                settings["saved_password"] = password
                logger.info("Saved credentials (Remember Me checked)")
            else:
                # Remove saved credentials if unchecked
                settings.pop("saved_username", None)
                settings.pop("saved_password", None)
                logger.info("Cleared saved credentials (Remember Me unchecked)")
            
            # Save settings
            with open(settings_path, 'w') as f:
                json.dump(settings, f, indent=2)
                
        except Exception as e:
            logger.error(f"Error saving credentials: {e}")
    
    def verify_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not username or not password:
            self.status_label.config(text="Please enter both username and password", fg="#ff4444")
            return
        
        # Check internet connection first
        if not check_internet_connection():
            self.status_label.config(text="❌ No internet connection", fg="#ff4444")
            messagebox.showerror("No Internet Connection", 
                               "Cannot connect to the internet.\n\nPlease check your network connection and try again.")
            return
        
        # Save credentials based on Remember Me checkbox
        self.save_credentials(username, password)
        
        # Show status (no progress bar)
        self.status_label.config(text="Verifying credentials...", fg=self.app.accent)
        self.login_button.config(state="disabled", text="Verifying...")
        self.frame.update()
        
        # Verify in background thread
        threading.Thread(target=self.verify_thread, args=(username, password), daemon=True).start()
    
    def verify_thread(self, username, password):
        """Verify credentials - ALWAYS HEADLESS"""
        driver = None
        try:
            print(f"\n=== VERIFYING CREDENTIALS ===")
            print(f"Username: {username}")
            
            # ALWAYS use headless for login verification (override user setting)
            from selenium.webdriver.chrome.options import Options
            chrome_options = Options()
            chrome_options.add_argument("--headless=new")  # Force headless
            chrome_options.add_argument("--incognito")
            chrome_options.add_argument("--force-device-scale-factor=0.75")
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            
            print("Chrome mode: Headless (invisible) - forced for login verification")
            
            print("Launching Chrome...")
            service = Service()
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(20)
            
            # Store driver in app for tool use
            self.app.driver = driver
            
            print("Navigating to login page...")
            driver.get(get_url("/auth/login", self))
            
            print("Waiting for page to load...")
            wait = WebDriverWait(driver, 10)
            time.sleep(3.0 / self.app.automation_speed)  # Page load
            
            print("Entering credentials...")
            actions = ActionChains(driver)
            actions.send_keys(Keys.TAB).perform()
            time.sleep(0.5 / self.app.automation_speed)  # Tab
            actions.send_keys(username).perform()
            time.sleep(0.1 / self.app.automation_speed)  # Text entry
            actions.send_keys(Keys.TAB).perform()
            time.sleep(0.5 / self.app.automation_speed)  # Tab
            actions.send_keys(password).perform()
            time.sleep(0.1 / self.app.automation_speed)  # Text entry
            
            print("Clicking LOGIN button...")
            try:
                login_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
                ))
                driver.execute_script("arguments[0].click();", login_button)
            except Exception as e:
                print(f"Failed to click login button: {e}")
                driver.quit()
                self.app.root.after(0, self.login_failed)
                return
            
            print("Waiting for navigation...")
            max_wait = 7
            start_time = time.time()
            success = False
            
            while (time.time() - start_time) < max_wait:
                current_url = driver.current_url
                print(f"Current URL: {current_url}")
                
                if "select-program" in current_url:
                    print("✅ Login successful!")
                    success = True
                    break
                
                if "login" not in current_url:
                    print("❌ Unexpected URL")
                    break
                
                time.sleep(0.5 / self.app.automation_speed)  # Check interval
            
            print("✓ Keeping Chrome open for tool use...")
            # Don't quit driver - keep it for tools to use
            # driver.quit()
            
            print(f"Verification result: {success}")
            print("=============================\n")
            
            if success:
                self.app.username = username
                self.app.password = password
                self.app.root.after(0, self.login_success)
            else:
                self.app.root.after(0, self.login_failed)
                
        except Exception as e:
            print(f"❌ Verification error: {e}")
            import traceback
            traceback.print_exc()
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            self.app.root.after(0, self.login_failed)
    
    def login_success(self):
        self.status_label.config(text="✓ Login successful!", fg="#28a745")
        self.login_button.config(text="✓ Success")
        self.frame.update()
        
        # Smooth transition
        self.app.root.after(500, self.app.show_main_menu)
    
    def login_failed(self):
        self.status_label.config(text="✗ Invalid credentials. Please try again.", fg="#ff4444")
        self.login_button.config(state="normal", text="Login →")
        self.progress_bar.place(relwidth=0)
    
    def destroy(self):
        self.frame.destroy()


# ======================================================================
# MAIN MENU SCREEN - Original UI Design
# ======================================================================

class AddUsersMenuScreen:
    """Submenu for Add Users tools"""
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.is_dark_mode = app.is_dark_mode
        self.bg_dark = app.bg_dark
        self.bg_card = app.bg_card
        self.accent = app.accent
        self.text_light = app.text_light
        self.text_dim = app.text_dim
        
        self._build_ui()
    
    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self.frame, bg=self.bg_dark, height=180)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button
        tk.Button(header_frame, 
                  text="← Back to Main Menu", 
                  command=self.app.show_main_menu,
                  bg=self.bg_card, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(x=30, y=25)
        
        # Title
        tk.Label(header_frame, 
                 text="Add Users", 
                 font=("Segoe UI", 36, "bold"),
                 bg=self.bg_dark,
                 fg=self.accent).pack(pady=(40, 0))
        
        tk.Label(header_frame, 
                 text="Select a user management system", 
                 font=("Segoe UI", 11),
                 bg=self.bg_dark,
                 fg=self.text_dim).pack(pady=(5, 0))
        
        # Content
        content_frame = tk.Frame(self.frame, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=40, pady=40)
        
        # Button container
        button_frame = tk.Frame(content_frame, bg=self.bg_dark)
        button_frame.pack(expand=True)
        
        button_config = {
            "font": ("Segoe UI", 14, "bold"),
            "relief": "flat",
            "cursor": "hand2",
            "width": 21,
            "height": 1,
            "bd": 0,
            "highlightthickness": 3,
            "highlightbackground": self.bg_dark,
            "activeforeground": "white"
        }
        
        # BBB Add User button
        bbb_color = "#4CAF50" if not self.is_dark_mode else "#3D8B40"
        tk.Button(button_frame, 
                  text="➕ BBB Add User", 
                  command=self.app.show_bbb_add_user,
                  bg=bbb_color,
                  fg="black",
                  activebackground="#45a049" if not self.is_dark_mode else "#2F6E33",
                  **button_config).pack(pady=12, padx=20)
        
        # SMS Add User button
        sms_color = "#FF9800" if not self.is_dark_mode else "#CC7A00"
        tk.Button(button_frame, 
                  text="➕ SMS Add User", 
                  command=self.app.show_sms_add_user,
                  bg=sms_color,
                  fg="black",
                  activebackground="#FB8C00",
                  **button_config).pack(pady=12, padx=20)
    
    def destroy(self):
        self.frame.destroy()


class BBBAddUserScreen:
    """BBB Add User - Upload spreadsheet and add users"""
    
    # Valid positions for BBB
    VALID_POSITIONS = [
        "Administrative Assistant",
        "Ambassador",
        "Cleaning Ambassador",
        "Client",
        "Custom Title",
        "Project Manager",
        "Outreach Coordinator",
        "Regional Director",
        "Regional Vice President",
        "Safety Ambassador",
        "Team Leader",
        "Vice President",
        "Hospitality Ambassador",
        "General Manager",
        "Operations Manager",
        "Operations Supervisor",
        "Outreach Ambassador"
    ]
    
    # Valid SMART roles for BBB (no custom options)
    VALID_SMART_ROLES = [
        "Ambassador",
        "Ambassador Elevated",
        "Customer Elevated",
        "Customer ONLY",
        "Divisional Vice President",
        "Full Access Role Test",
        "Managment",
        "MBTA Admin",
        "Outreach Coordinator",
        "Regional Vice President",
        "Supervisor",
        "System Administrator",
        "Team Lead"
    ]
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.root = app.root
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.is_dark_mode = app.is_dark_mode
        self.bg_dark = app.bg_dark
        self.bg_card = app.bg_card
        self.accent = app.accent
        self.text_light = app.text_light
        self.text_dim = app.text_dim
        self.username = app.username
        self.password = app.password
        self.headless_mode = app.headless_mode
        
        self.users_data = []  # Will hold parsed user data
        
        self._build_ui()
    
    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self.frame, bg=self.bg_dark, height=120)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button
        tk.Button(header_frame, 
                  text="← Back", 
                  command=self.app.show_add_users_menu,
                  bg=self.bg_card, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(x=30, y=15)
        
        # Title
        tk.Label(header_frame, 
                 text="BBB Add User", 
                 font=("Segoe UI", 24, "bold"),
                 bg=self.bg_dark,
                 fg=self.accent).pack(pady=(20, 0))
        
        tk.Label(header_frame, 
                 text="Upload spreadsheet to add users", 
                 font=("Segoe UI", 11),
                 bg=self.bg_dark,
                 fg=self.text_dim).pack(pady=(5, 0))
        
        # Content area
        content_frame = tk.Frame(self.frame, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=40, pady=20)
        
        # Top controls
        controls_frame = tk.Frame(content_frame, bg=self.bg_dark)
        controls_frame.pack(fill="x", pady=(0, 15))
        
        # Select Spreadsheet button
        tk.Button(controls_frame,
                 text="📂 Select Spreadsheet",
                 command=self.select_spreadsheet,
                 bg=self.accent,
                 fg="white",
                 font=("Segoe UI", 11, "bold"),
                 relief="flat",
                 cursor="hand2",
                 padx=20,
                 pady=10).pack(side="left", padx=(0, 15))
        
        # File label
        self.file_label = tk.Label(controls_frame,
                                   text="No file selected",
                                   font=("Segoe UI", 10),
                                   bg=self.bg_dark,
                                   fg=self.text_dim)
        self.file_label.pack(side="left")
        
        # Add Users button (initially disabled)
        self.add_users_button = tk.Button(controls_frame,
                                          text="➕ Add Users",
                                          command=self.start_add_users,
                                          bg="#4CAF50",
                                          fg="white",
                                          font=("Segoe UI", 11, "bold"),
                                          relief="flat",
                                          cursor="hand2",
                                          padx=20,
                                          pady=10,
                                          state="disabled")
        self.add_users_button.pack(side="right")
        
        # Table frame
        table_frame = tk.Frame(content_frame, bg=self.bg_card, relief="solid", borderwidth=1)
        table_frame.pack(fill="both", expand=True)
        
        # Table header
        tk.Label(table_frame,
                text="Users to Add",
                font=("Segoe UI", 12, "bold"),
                bg=self.bg_card,
                fg=self.accent,
                anchor="w",
                padx=15,
                pady=10).pack(fill="x")
        
        # Table with scrollbar
        table_container = tk.Frame(table_frame, bg=self.bg_card)
        table_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Scrollable frame setup
        canvas = tk.Canvas(table_container, bg=self.bg_card, highlightthickness=0)
        scrollbar = tk.Scrollbar(table_container, orient="vertical", command=canvas.yview)
        self.table_content = tk.Frame(canvas, bg=self.bg_card)
        
        self.table_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.table_content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Initial empty state message
        self.empty_label = tk.Label(self.table_content,
                                    text="No users loaded.\nSelect a spreadsheet to begin.",
                                    font=("Segoe UI", 11),
                                    bg=self.bg_card,
                                    fg=self.text_dim,
                                    pady=50)
        self.empty_label.pack()
    
    def select_spreadsheet(self):
        """Open file dialog to select spreadsheet"""
        from tkinter import filedialog
        
        filepath = filedialog.askopenfilename(
            title="Select Spreadsheet",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if filepath:
            logger.info(f"Selected spreadsheet: {filepath}")
            self.file_label.config(text=f"File: {filepath.split('/')[-1]}")
            self.parse_spreadsheet(filepath)
    
    def parse_spreadsheet(self, filepath):
        """Parse the spreadsheet and populate the table"""
        import openpyxl
        
        try:
            logger.info("Parsing spreadsheet...")
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Expected columns from CVG format:
            # Col 1: Location Name
            # Col 2: Location No.
            # Col 3: Legal First Name
            # Col 4: Legal Last Name
            # Col 5: Position
            # Col 6: SMART Role
            # Col 7: Seniority Date
            # Col 8: Work Contact: Work Email
            # Col 9: Username
            # Col 10: Password
            
            self.users_data = []
            
            # Skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if row has data (first name exists)
                if row[2]:  # Legal First Name
                    user = {
                        "location_name": str(row[0]).strip() if row[0] else "",
                        "location_no": str(row[1]).strip() if row[1] else "",
                        "first_name": str(row[2]).strip() if row[2] else "",
                        "last_name": str(row[3]).strip() if row[3] else "",
                        "position": str(row[4]).strip() if row[4] else "",
                        "smart_role": str(row[5]).strip() if row[5] else "",
                        "seniority_date": str(row[6]).strip() if row[6] else "",
                        "email": str(row[7]).strip() if row[7] else "",
                        "username": str(row[8]).strip() if row[8] else "",
                        "password": str(row[9]).strip() if row[9] else ""
                    }
                    self.users_data.append(user)
            
            logger.info(f"Parsed {len(self.users_data)} users from spreadsheet")
            
            if len(self.users_data) == 0:
                from tkinter import messagebox
                messagebox.showwarning("No Data", "No user data found in spreadsheet.\n\nPlease ensure the spreadsheet has data starting from row 2.")
                return
            
            self.display_users_table()
            
            # Enable Add Users button
            self.add_users_button.config(state="normal")
            
        except Exception as e:
            logger.error(f"Error parsing spreadsheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            from tkinter import messagebox
            messagebox.showerror("Error", f"Failed to parse spreadsheet:\n{str(e)}")
    
    def display_users_table(self):
        """Display users in an editable table with checkboxes"""
        # Clear existing content
        for widget in self.table_content.winfo_children():
            widget.destroy()
        
        if not self.users_data:
            self.empty_label.pack()
            return
        
        # Create frame for table
        table_frame = tk.Frame(self.table_content, bg=self.bg_card)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Headers
        headers = ["✓", "First Name", "Last Name", "Position", "SMART Role", "Email", "Username", "Password", "Location Name", "Location No"]
        col_widths = [3, 12, 12, 15, 15, 20, 12, 12, 15, 10]
        
        # Header row
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            tk.Label(table_frame,
                    text=header,
                    font=("Segoe UI", 9, "bold"),
                    bg=self.bg_dark if self.is_dark_mode else "#e0e0e0",
                    fg=self.accent,
                    width=width,
                    borderwidth=1,
                    relief="solid",
                    anchor="w",
                    padx=5,
                    pady=5).grid(row=0, column=col, sticky="ew")
        
        # Data rows with editable entries
        for idx, user in enumerate(self.users_data):
            row_num = idx + 1
            
            # Add checkbox variable to user data
            user['selected'] = tk.BooleanVar(value=True)  # Selected by default
            
            # Checkbox
            chk = tk.Checkbutton(table_frame,
                                variable=user['selected'],
                                bg=self.bg_card,
                                fg=self.text_light,
                                selectcolor=self.bg_dark,
                                activebackground=self.bg_card,
                                activeforeground=self.text_light,
                                cursor="hand2")
            chk.grid(row=row_num, column=0, sticky="ew")
            
            # Store entry widgets in user dict for later retrieval
            user['entries'] = {}
            
            # Editable fields
            fields = [
                ('first_name', user['first_name']),
                ('last_name', user['last_name']),
                ('position', user['position']),
                ('smart_role', user['smart_role']),
                ('email', user['email']),
                ('username', user['username']),
                ('password', user['password']),
                ('location_name', user['location_name']),
                ('location_no', user['location_no'])
            ]
            
            for col, (field_name, field_value) in enumerate(fields, 1):
                entry = tk.Entry(table_frame,
                                font=("Segoe UI", 9),
                                bg=self.bg_card,
                                fg=self.text_light,
                                width=col_widths[col],
                                borderwidth=1,
                                relief="solid",
                                insertbackground=self.text_light)
                entry.insert(0, field_value)
                entry.grid(row=row_num, column=col, sticky="ew", padx=1, pady=1)
                
                # Store entry widget reference
                user['entries'][field_name] = entry
        
        logger.info(f"Displayed {len(self.users_data)} users in editable table")
        
        # Add select/deselect all buttons at bottom
        button_frame = tk.Frame(self.table_content, bg=self.bg_card)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(button_frame,
                 text="✓ Select All",
                 command=self.select_all_users,
                 bg=self.bg_dark,
                 fg=self.text_light,
                 font=("Segoe UI", 9),
                 relief="flat",
                 cursor="hand2",
                 padx=10,
                 pady=5).pack(side="left", padx=5)
        
        tk.Button(button_frame,
                 text="✗ Deselect All",
                 command=self.deselect_all_users,
                 bg=self.bg_dark,
                 fg=self.text_light,
                 font=("Segoe UI", 9),
                 relief="flat",
                 cursor="hand2",
                 padx=10,
                 pady=5).pack(side="left", padx=5)
        
        # Show count
        selected_count = sum(1 for user in self.users_data if user['selected'].get())
        self.count_label = tk.Label(button_frame,
                                     text=f"{selected_count} of {len(self.users_data)} users selected",
                                     font=("Segoe UI", 9),
                                     bg=self.bg_card,
                                     fg=self.text_dim)
        self.count_label.pack(side="right", padx=10)
        
        # Update count when checkboxes change
        for user in self.users_data:
            user['selected'].trace('w', lambda *args: self.update_count())
    
    def select_all_users(self):
        """Select all users"""
        for user in self.users_data:
            user['selected'].set(True)
    
    def deselect_all_users(self):
        """Deselect all users"""
        for user in self.users_data:
            user['selected'].set(False)
    
    def update_count(self):
        """Update the selected count label"""
        if hasattr(self, 'count_label'):
            selected_count = sum(1 for user in self.users_data if user['selected'].get())
            self.count_label.config(text=f"{selected_count} of {len(self.users_data)} users selected")
    
    def get_selected_users(self):
        """Get list of selected users with updated values from entry fields"""
        selected_users = []
        
        for user in self.users_data:
            if user['selected'].get():
                # Get updated values from entry widgets
                updated_user = {
                    'first_name': user['entries']['first_name'].get(),
                    'last_name': user['entries']['last_name'].get(),
                    'position': user['entries']['position'].get(),
                    'smart_role': user['entries']['smart_role'].get(),
                    'email': user['entries']['email'].get(),
                    'username': user['entries']['username'].get(),
                    'password': user['entries']['password'].get(),
                    'location_name': user['entries']['location_name'].get(),
                    'location_no': user['entries']['location_no'].get()
                }
                selected_users.append(updated_user)
        
        return selected_users
    
    def start_add_users(self):
        """Start the add users process"""
        if not self.users_data:
            from tkinter import messagebox
            messagebox.showwarning("No Users", "No users to add. Please select a spreadsheet first.")
            return
        
        # Get selected users with their edited values
        selected_users = self.get_selected_users()
        
        if not selected_users:
            from tkinter import messagebox
            messagebox.showwarning("No Users Selected", "Please select at least one user to add.")
            return
        
        logger.info(f"Starting add users process for {len(selected_users)} selected users")
        logger.info(f"Environment: {self.app.environment}")
        
        # Store selected users for the automation
        self.selected_users = selected_users
        
        # Show warning if not in headless mode
        if not self.app.headless_mode:
            self.app.show_dark_messagebox("Non-Headless Mode Active",
                                         "⚠️ Chrome is running in visible mode.\n\n"
                                         "Please DO NOT:\n"
                                         "• Minimize the Chrome window\n"
                                         "• Click away from Chrome\n"
                                         "• Close Chrome manually\n\n"
                                         "Keep Chrome visible during the add user process.",
                                         "warning")
        
        # Navigate to users page and click Add User button
        import threading
        threading.Thread(target=self.run_add_users, daemon=True).start()
    
    def run_add_users(self):
        """Run the add users automation"""
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        
        driver = None
        
        try:
            logger.info("="*60)
            logger.info("BBB ADD USERS - Starting")
            logger.info(f"Environment: {self.app.environment}")
            logger.info(f"Users to add: {len(self.selected_users)}")
            logger.info("="*60)
            
            # Launch Chrome
            chrome_options = self.app.get_chrome_options()
            if not self.headless_mode:
                chrome_options.add_argument("--start-maximized")
            
            service = Service()
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(20)
            
            wait = WebDriverWait(driver, 10)
            actions = ActionChains(driver)
            
            # Login
            logger.info("Logging in...")
            driver.get(get_url("/auth/login", self.app))
            self.app.automation_sleep(3)  # Page load
            
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys(self.username).perform()
            self.app.automation_sleep(0.1)  # Text entry
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys(self.password).perform()
            self.app.automation_sleep(0.1)  # Text entry
            
            login_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
            ))
            driver.execute_script("arguments[0].click();", login_button)
            self.app.automation_sleep(3)  # Page load
            
            # Select program
            logger.info("Selecting program...")
            self.app.automation_sleep(1)  # Wait before program selection
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys("0000").perform()
            self.app.automation_sleep(0.1)  # Text entry
            actions.send_keys(Keys.ARROW_DOWN).perform()
            self.app.automation_sleep(0.5)  # Keypress
            actions.send_keys(Keys.ENTER).perform()
            self.app.automation_sleep(0.5)  # Keypress
            
            submit_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
            ))
            driver.execute_script("arguments[0].click();", submit_button)
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ Logged in")
            
            # Navigate to users page
            logger.info("Navigating to users page...")
            driver.get(get_url("/main/users", self.app))
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ On users page")
            
            # Click Add User button
            logger.info("Looking for Add User button...")
            
            # Try the CSS selector
            add_user_button = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#root > div.collapsed.AdminLayout_admin_layout_container__8z6D1 > div.AdminLayout_admin_body_container__qHWPk > div > div:nth-child(1) > div.d-flex.justify-content-between.align-items-center.mb-2.bg-white.p-3.undefined.undefined.undefined > div.d-flex.justify-content-end.align-items-center > div > button")
            ))
            
            logger.info("✓ Found Add User button")
            driver.execute_script("arguments[0].click();", add_user_button)
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ Clicked Add User button")
            logger.info(f"Current URL: {driver.current_url}")
            
            # Tab 9 times to navigate on the add user form
            logger.info("Navigating form with 9 tabs...")
            for i in range(9):
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)  # Tab
            
            logger.info("✓ Navigated to form field")
            
            # Process first user (testing with one user)
            if self.selected_users:
                user = self.selected_users[0]
                logger.info(f"Adding user: {user['first_name']} {user['last_name']}")
                
                # Type first name
                actions.send_keys(user['first_name']).perform()
                self.app.automation_sleep(0.1)  # Text
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type password (from spreadsheet)
                actions.send_keys(user['password']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Space
                actions.send_keys(Keys.SPACE).perform()
                self.app.automation_sleep(0.5)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type Last Name
                actions.send_keys(user['last_name']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab (3x)
                for _ in range(3):
                    actions.send_keys(Keys.TAB).perform()
                    self.app.automation_sleep(0.5)
                
                # Space (2x)
                for _ in range(2):
                    actions.send_keys(Keys.SPACE).perform()
                    self.app.automation_sleep(0.5)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type username
                actions.send_keys(user['username']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Space
                actions.send_keys(Keys.SPACE).perform()
                self.app.automation_sleep(0.5)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type email
                actions.send_keys(user['email']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type position
                actions.send_keys(user['position']).perform()
                self.app.automation_sleep(0.1)
                
                # Enter
                actions.send_keys(Keys.ENTER).perform()
                self.app.automation_sleep(0.5)
                
                # Escape
                actions.send_keys(Keys.ESCAPE).perform()
                self.app.automation_sleep(0.5)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type smart role
                actions.send_keys(user['smart_role']).perform()
                self.app.automation_sleep(0.1)
                
                # Enter
                actions.send_keys(Keys.ENTER).perform()
                self.app.automation_sleep(0.5)
                
                # Tab
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)
                
                # Type password again
                actions.send_keys(user['password']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab (2x)
                for _ in range(2):
                    actions.send_keys(Keys.TAB).perform()
                    self.app.automation_sleep(0.5)
                
                # Type location number
                actions.send_keys(user['location_no']).perform()
                self.app.automation_sleep(0.1)
                
                # Tab (4x)
                for _ in range(4):
                    actions.send_keys(Keys.TAB).perform()
                    self.app.automation_sleep(0.5)
                
                # Space
                actions.send_keys(Keys.SPACE).perform()
                self.app.automation_sleep(1)  # Wait after checkbox
                
                # Click Save button
                logger.info("Clicking Save button...")
                save_button = wait.until(EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "#root > div.collapsed.AdminLayout_admin_layout_container__8z6D1 > div.AdminLayout_admin_body_container__qHWPk > div > div:nth-child(4) > div.content-container > div.d-flex.justify-content-end.mt-5.mb-4 > div:nth-child(2) > div > button")
                ))
                driver.execute_script("arguments[0].click();", save_button)
                self.app.automation_sleep(3)  # Page load
                
                logger.info(f"✓ User {user['first_name']} {user['last_name']} added successfully!")
            
            # Keep browser open for now (testing)
            logger.info("Browser will stay open for testing...")
            logger.info("Automation paused - close browser manually to continue")
            
        except Exception as e:
            logger.error(f"Error during add users: {e}")
            import traceback
            logger.error(traceback.format_exc())
            
        finally:
            # Don't close browser for now - for testing
            pass


class SMSAddUserScreen:
    """SMS Add User - Upload spreadsheet and add users"""
    
    # Valid positions for SMS
    VALID_POSITIONS = [
        "Administrative Assistant",
        "cleaner",
        "cleaning ambassador",
        "client",
        "custom Title",
        "general Manager",
        "Hospitality ambassador",
        "operations manager",
        "regional vice president",
        "safety ambassador",
        "team leader",
        "vice president",
        "project manager",
        "outreach coordinator",
        "outreach ambassador",
        "operation supervisor",
        "regional director"
    ]
    
    # Valid SMART roles for SMS (same as BBB)
    VALID_SMART_ROLES = [
        "Ambassador",
        "Ambassador Elevated",
        "Customer Elevated",
        "Customer ONLY",
        "Divisional Vice President",
        "Full Access Role Test",
        "Managment",
        "MBTA Admin",
        "Outreach Coordinator",
        "Regional Vice President",
        "Supervisor",
        "System Administrator",
        "Team Lead"
    ]
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.root = app.root
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.is_dark_mode = app.is_dark_mode
        self.bg_dark = app.bg_dark
        self.bg_card = app.bg_card
        self.accent = app.accent
        self.text_light = app.text_light
        self.text_dim = app.text_dim
        self.username = app.username
        self.password = app.password
        self.headless_mode = app.headless_mode
        
        self.users_data = []  # Will hold parsed user data
        
        self._build_ui()
    
    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self.frame, bg=self.bg_dark, height=120)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button
        tk.Button(header_frame, 
                  text="← Back", 
                  command=self.app.show_add_users_menu,
                  bg=self.bg_card, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(x=30, y=15)
        
        # Title
        tk.Label(header_frame, 
                 text="SMS Add User", 
                 font=("Segoe UI", 24, "bold"),
                 bg=self.bg_dark,
                 fg="#FF9800").pack(pady=(20, 0))  # Orange accent for SMS
        
        tk.Label(header_frame, 
                 text="Upload spreadsheet to add users", 
                 font=("Segoe UI", 11),
                 bg=self.bg_dark,
                 fg=self.text_dim).pack(pady=(5, 0))
        
        # Content area
        content_frame = tk.Frame(self.frame, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=40, pady=20)
        
        # Top controls
        controls_frame = tk.Frame(content_frame, bg=self.bg_dark)
        controls_frame.pack(fill="x", pady=(0, 15))
        
        # Select Spreadsheet button
        tk.Button(controls_frame,
                 text="📂 Select Spreadsheet",
                 command=self.select_spreadsheet,
                 bg="#FF9800",  # Orange for SMS
                 fg="white",
                 font=("Segoe UI", 11, "bold"),
                 relief="flat",
                 cursor="hand2",
                 padx=20,
                 pady=10).pack(side="left", padx=(0, 15))
        
        # File label
        self.file_label = tk.Label(controls_frame,
                                   text="No file selected",
                                   font=("Segoe UI", 10),
                                   bg=self.bg_dark,
                                   fg=self.text_dim)
        self.file_label.pack(side="left")
        
        # Add Users button (initially disabled)
        self.add_users_button = tk.Button(controls_frame,
                                          text="➕ Add Users",
                                          command=self.start_add_users,
                                          bg="#4CAF50",
                                          fg="white",
                                          font=("Segoe UI", 11, "bold"),
                                          relief="flat",
                                          cursor="hand2",
                                          padx=20,
                                          pady=10,
                                          state="disabled")
        self.add_users_button.pack(side="right")
        
        # Table frame
        table_frame = tk.Frame(content_frame, bg=self.bg_card, relief="solid", borderwidth=1)
        table_frame.pack(fill="both", expand=True)
        
        # Table header
        tk.Label(table_frame,
                text="Users to Add",
                font=("Segoe UI", 12, "bold"),
                bg=self.bg_card,
                fg="#FF9800",  # Orange for SMS
                anchor="w",
                padx=15,
                pady=10).pack(fill="x")
        
        # Table with scrollbar
        table_container = tk.Frame(table_frame, bg=self.bg_card)
        table_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Scrollable frame setup
        canvas = tk.Canvas(table_container, bg=self.bg_card, highlightthickness=0)
        scrollbar = tk.Scrollbar(table_container, orient="vertical", command=canvas.yview)
        self.table_content = tk.Frame(canvas, bg=self.bg_card)
        
        self.table_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.table_content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Initial empty state message
        self.empty_label = tk.Label(self.table_content,
                                    text="No users loaded.\nSelect a spreadsheet to begin.",
                                    font=("Segoe UI", 11),
                                    bg=self.bg_card,
                                    fg=self.text_dim,
                                    pady=50)
        self.empty_label.pack()
    
    def select_spreadsheet(self):
        """Open file dialog to select spreadsheet"""
        from tkinter import filedialog
        
        filepath = filedialog.askopenfilename(
            title="Select Spreadsheet",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if filepath:
            logger.info(f"Selected spreadsheet: {filepath}")
            self.file_label.config(text=f"File: {filepath.split('/')[-1]}")
            self.parse_spreadsheet(filepath)
    
    def parse_spreadsheet(self, filepath):
        """Parse the spreadsheet and populate the table"""
        import openpyxl
        
        try:
            logger.info("Parsing spreadsheet...")
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            self.users_data = []
            
            # Skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[2]:  # Legal First Name
                    user = {
                        "location_name": str(row[0]).strip() if row[0] else "",
                        "location_no": str(row[1]).strip() if row[1] else "",
                        "first_name": str(row[2]).strip() if row[2] else "",
                        "last_name": str(row[3]).strip() if row[3] else "",
                        "position": str(row[4]).strip() if row[4] else "",
                        "smart_role": str(row[5]).strip() if row[5] else "",
                        "seniority_date": str(row[6]).strip() if row[6] else "",
                        "email": str(row[7]).strip() if row[7] else "",
                        "username": str(row[8]).strip() if row[8] else "",
                        "password": str(row[9]).strip() if row[9] else ""
                    }
                    self.users_data.append(user)
            
            logger.info(f"Parsed {len(self.users_data)} users from spreadsheet")
            
            if len(self.users_data) == 0:
                from tkinter import messagebox
                messagebox.showwarning("No Data", "No user data found in spreadsheet.\n\nPlease ensure the spreadsheet has data starting from row 2.")
                return
            
            self.display_users_table()
            self.add_users_button.config(state="normal")
            
        except Exception as e:
            logger.error(f"Error parsing spreadsheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            from tkinter import messagebox
            messagebox.showerror("Error", f"Failed to parse spreadsheet:\n{str(e)}")
    
    def display_users_table(self):
        """Display users in an editable table with checkboxes"""
        # Clear existing content
        for widget in self.table_content.winfo_children():
            widget.destroy()
        
        if not self.users_data:
            self.empty_label.pack()
            return
        
        # Create frame for table
        table_frame = tk.Frame(self.table_content, bg=self.bg_card)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Headers
        headers = ["✓", "First Name", "Last Name", "Position", "SMART Role", "Email", "Username", "Password", "Location Name", "Location No"]
        col_widths = [3, 12, 12, 15, 15, 20, 12, 12, 15, 10]
        
        # Header row
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            tk.Label(table_frame,
                    text=header,
                    font=("Segoe UI", 9, "bold"),
                    bg=self.bg_dark if self.is_dark_mode else "#e0e0e0",
                    fg="#FF9800",  # Orange for SMS
                    width=width,
                    borderwidth=1,
                    relief="solid",
                    anchor="w",
                    padx=5,
                    pady=5).grid(row=0, column=col, sticky="ew")
        
        # Data rows with editable entries
        for idx, user in enumerate(self.users_data):
            row_num = idx + 1
            user['selected'] = tk.BooleanVar(value=True)
            
            # Checkbox
            chk = tk.Checkbutton(table_frame,
                                variable=user['selected'],
                                bg=self.bg_card,
                                fg=self.text_light,
                                selectcolor=self.bg_dark,
                                activebackground=self.bg_card,
                                activeforeground=self.text_light,
                                cursor="hand2")
            chk.grid(row=row_num, column=0, sticky="ew")
            
            user['entries'] = {}
            
            fields = [
                ('first_name', user['first_name']),
                ('last_name', user['last_name']),
                ('position', user['position']),
                ('smart_role', user['smart_role']),
                ('email', user['email']),
                ('username', user['username']),
                ('password', user['password']),
                ('location_name', user['location_name']),
                ('location_no', user['location_no'])
            ]
            
            for col, (field_name, field_value) in enumerate(fields, 1):
                entry = tk.Entry(table_frame,
                                font=("Segoe UI", 9),
                                bg=self.bg_card,
                                fg=self.text_light,
                                width=col_widths[col],
                                borderwidth=1,
                                relief="solid",
                                insertbackground=self.text_light)
                entry.insert(0, field_value)
                entry.grid(row=row_num, column=col, sticky="ew", padx=1, pady=1)
                user['entries'][field_name] = entry
        
        logger.info(f"Displayed {len(self.users_data)} users in editable table")
        
        # Add select/deselect all buttons at bottom
        button_frame = tk.Frame(self.table_content, bg=self.bg_card)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(button_frame,
                 text="✓ Select All",
                 command=self.select_all_users,
                 bg=self.bg_dark,
                 fg=self.text_light,
                 font=("Segoe UI", 9),
                 relief="flat",
                 cursor="hand2",
                 padx=10,
                 pady=5).pack(side="left", padx=5)
        
        tk.Button(button_frame,
                 text="✗ Deselect All",
                 command=self.deselect_all_users,
                 bg=self.bg_dark,
                 fg=self.text_light,
                 font=("Segoe UI", 9),
                 relief="flat",
                 cursor="hand2",
                 padx=10,
                 pady=5).pack(side="left", padx=5)
        
        selected_count = sum(1 for user in self.users_data if user['selected'].get())
        self.count_label = tk.Label(button_frame,
                                     text=f"{selected_count} of {len(self.users_data)} users selected",
                                     font=("Segoe UI", 9),
                                     bg=self.bg_card,
                                     fg=self.text_dim)
        self.count_label.pack(side="right", padx=10)
        
        for user in self.users_data:
            user['selected'].trace('w', lambda *args: self.update_count())
    
    def select_all_users(self):
        for user in self.users_data:
            user['selected'].set(True)
    
    def deselect_all_users(self):
        for user in self.users_data:
            user['selected'].set(False)
    
    def update_count(self):
        if hasattr(self, 'count_label'):
            selected_count = sum(1 for user in self.users_data if user['selected'].get())
            self.count_label.config(text=f"{selected_count} of {len(self.users_data)} users selected")
    
    def get_selected_users(self):
        """Get list of selected users with updated values from entry fields"""
        selected_users = []
        
        for user in self.users_data:
            if user['selected'].get():
                updated_user = {
                    'first_name': user['entries']['first_name'].get(),
                    'last_name': user['entries']['last_name'].get(),
                    'position': user['entries']['position'].get(),
                    'smart_role': user['entries']['smart_role'].get(),
                    'email': user['entries']['email'].get(),
                    'username': user['entries']['username'].get(),
                    'password': user['entries']['password'].get(),
                    'location_name': user['entries']['location_name'].get(),
                    'location_no': user['entries']['location_no'].get()
                }
                selected_users.append(updated_user)
        
        return selected_users
    
    def start_add_users(self):
        """Start the add users process"""
        if not self.users_data:
            from tkinter import messagebox
            messagebox.showwarning("No Users", "No users to add. Please select a spreadsheet first.")
            return
        
        selected_users = self.get_selected_users()
        
        if not selected_users:
            from tkinter import messagebox
            messagebox.showwarning("No Users Selected", "Please select at least one user to add.")
            return
        
        logger.info(f"Starting SMS add users process for {len(selected_users)} selected users")
        
        self.selected_users = selected_users
        
        # Show warning if not in headless mode
        if not self.app.headless_mode:
            self.app.show_dark_messagebox("Non-Headless Mode Active",
                                         "⚠️ Chrome is running in visible mode.\n\n"
                                         "Please DO NOT:\n"
                                         "• Minimize the Chrome window\n"
                                         "• Click away from Chrome\n"
                                         "• Close Chrome manually\n\n"
                                         "Keep Chrome visible during the add user process.",
                                         "warning")
        
        import threading
        threading.Thread(target=self.run_add_users, daemon=True).start()
    
    def run_add_users(self):
        """Run the SMS add users automation"""
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        
        driver = None
        
        try:
            logger.info("="*60)
            logger.info("SMS ADD USERS - Starting")
            logger.info(f"Users to add: {len(self.selected_users)}")
            logger.info("="*60)
            
            # Launch Chrome
            chrome_options = self.app.get_chrome_options()
            if not self.headless_mode:
                chrome_options.add_argument("--start-maximized")
            
            service = Service()
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(20)
            
            wait = WebDriverWait(driver, 10)
            actions = ActionChains(driver)
            
            # Login to SMS
            logger.info("Logging in to SMS...")
            driver.get("https://fe.smsclean.com/auth/login")
            self.app.automation_sleep(3)  # Page load
            
            # Username
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys(self.username).perform()
            self.app.automation_sleep(0.1)  # Text entry
            
            # Password
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys(self.password).perform()
            self.app.automation_sleep(0.1)  # Text entry
            
            # Click login
            login_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'LOGIN')]")
            ))
            driver.execute_script("arguments[0].click();", login_button)
            self.app.automation_sleep(3)  # Page load
            
            # Select program - z-demo (same as BBB but z-demo instead of 0000)
            logger.info("Selecting program...")
            self.app.automation_sleep(1)  # Wait before program selection
            actions.send_keys(Keys.TAB).perform()
            self.app.automation_sleep(0.5)  # Tab
            actions.send_keys("z-demo").perform()
            self.app.automation_sleep(0.1)  # Text entry
            actions.send_keys(Keys.ARROW_DOWN).perform()
            self.app.automation_sleep(0.5)  # Keypress
            actions.send_keys(Keys.ENTER).perform()
            self.app.automation_sleep(0.5)  # Keypress
            
            submit_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'login-input') and contains(text(), 'SUBMIT')]")
            ))
            driver.execute_script("arguments[0].click();", submit_button)
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ Logged in to SMS")
            
            # Verify dashboard
            logger.info("Verifying dashboard...")
            expected_dashboard = "https://fe.smsclean.com/main/dashboard?tab=1"
            if expected_dashboard in driver.current_url:
                logger.info(f"✓ Confirmed on dashboard: {driver.current_url}")
            else:
                logger.warning(f"⚠ Not on expected dashboard. Current: {driver.current_url}")
            
            # Navigate to users page
            logger.info("Navigating to users page...")
            driver.get("https://fe.smsclean.com/main/users")
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ On users page")
            
            # Click Add User button
            logger.info("Looking for Add User button...")
            
            add_user_button = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#root > div.collapsed.AdminLayout_admin_layout_container__8z6D1 > div.AdminLayout_admin_body_container__qHWPk > div > div:nth-child(1) > div.d-flex.justify-content-between.align-items-center.mb-2.bg-white.p-3.undefined.undefined.undefined > div.d-flex.justify-content-end.align-items-center > div > button")
            ))
            
            logger.info("✓ Found Add User button")
            driver.execute_script("arguments[0].click();", add_user_button)
            self.app.automation_sleep(3)  # Page load
            
            logger.info("✓ Clicked Add User button")
            logger.info(f"Current URL: {driver.current_url}")
            
            # Tab 9 times to navigate on the add user form
            logger.info("Navigating form with 9 tabs...")
            for i in range(9):
                actions.send_keys(Keys.TAB).perform()
                self.app.automation_sleep(0.5)  # Tab
            
            logger.info("✓ Navigated to form field")
            
            # Keep browser open for testing
            logger.info("Browser will stay open for testing...")
            logger.info("Automation paused - close browser manually to continue")
            
        except Exception as e:
            logger.error(f"Error during SMS add users: {e}")
            import traceback
            logger.error(traceback.format_exc())
            
        finally:
            # Don't close browser for now - for testing
            pass


class StandardsScansMenuScreen:
    """Submenu for Standards Scans tools"""
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.is_dark_mode = app.is_dark_mode
        self.bg_dark = app.bg_dark
        self.bg_card = app.bg_card
        self.accent = app.accent
        self.text_light = app.text_light
        self.text_dim = app.text_dim
        
        self._build_ui()
    
    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self.frame, bg=self.bg_dark, height=180)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Back button
        tk.Button(header_frame, 
                  text="← Back to Main Menu", 
                  command=self.app.show_main_menu,
                  bg=self.bg_card, 
                  fg=self.text_light,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat",
                  cursor="hand2",
                  padx=20,
                  pady=8,
                  highlightthickness=0).place(x=30, y=25)
        
        # Title
        tk.Label(header_frame, 
                 text="Standards Scans", 
                 font=("Segoe UI", 36, "bold"),
                 bg=self.bg_dark,
                 fg=self.accent).pack(pady=(40, 0))
        
        tk.Label(header_frame, 
                 text="Select a standards scanning tool", 
                 font=("Segoe UI", 11),
                 bg=self.bg_dark,
                 fg=self.text_dim).pack(pady=(5, 0))
        
        # Content
        content_frame = tk.Frame(self.frame, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=40, pady=40)
        
        # Button container
        button_frame = tk.Frame(content_frame, bg=self.bg_dark)
        button_frame.pack(expand=True)
        
        button_config = {
            "font": ("Segoe UI", 14, "bold"),
            "relief": "flat",
            "cursor": "hand2",
            "width": 21,
            "height": 1,
            "bd": 0,
            "highlightthickness": 3,
            "highlightbackground": self.bg_dark,
            "activeforeground": "white"
        }
        
        # User Standards Scan button
        standards_color = "#FFD951" if not self.is_dark_mode else "#CCA800"
        tk.Button(button_frame, 
                  text="📋 Users Scan", 
                  command=self.app.show_user_standards,
                  bg=standards_color,
                  fg="black",
                  activebackground="#FFC936" if not self.is_dark_mode else "#B38600",
                  **button_config).pack(pady=12, padx=20)
        
        # Equipment Standards Scan button
        equipment_color = "#0052CC" if self.is_dark_mode else "#0066ff"
        tk.Button(button_frame, 
                  text="🔧 Equipment Scan", 
                  command=self.app.show_equipment_standards,
                  bg=equipment_color,
                  fg="black",
                  activebackground="#003d99" if self.is_dark_mode else "#0052CC",
                  **button_config).pack(pady=12, padx=20)
    
    def destroy(self):
        self.frame.destroy()


class MainMenuScreen:
    """Main menu with original design - circular quick links, rounded buttons"""
    
    def __init__(self, parent_frame, app):
        self.app = app
        self.frame = tk.Frame(parent_frame, bg=app.bg_dark)
        self.frame.pack(fill="both", expand=True)
        
        self.is_dark_mode = app.is_dark_mode
        self.bg_dark = app.bg_dark
        self.bg_card = app.bg_card
        self.accent = app.accent
        self.text_light = app.text_light
        self.text_dim = app.text_dim
        
        self._build_ui()
        
        # Check internet connection after UI is built
        self.frame.after(500, self.check_internet_on_startup)
    
    def check_internet_on_startup(self):
        """Check internet connection and show warning if offline"""
        if not check_internet_connection():
            from tkinter import messagebox
            messagebox.showwarning("No Internet Connection",
                                 "⚠️ No internet connection detected.\n\n"
                                 "You won't be able to use automation features until you're online.\n\n"
                                 "Please check your network connection.")
    
    def _build_ui(self):
        """Build the main menu UI with modern circular quick links"""
        # Header
        header_frame = tk.Frame(self.frame, bg=self.bg_dark, height=180)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        # Quick Links section (top left) - Simple text links
        links_frame = tk.Frame(header_frame, bg=self.bg_dark)
        links_frame.place(x=30, y=25)
        
        # Header
        header_label = tk.Label(links_frame, text="Quick Links", 
                               font=("Segoe UI", 12, "bold"),
                               bg=self.bg_dark, fg=self.accent)
        header_label.pack(anchor="w", pady=(0, 10))
        
        # Create text link
        def create_text_link(parent, text, url):
            link = tk.Label(parent, text=text, font=("Segoe UI", 10, "underline"),
                          bg=self.bg_dark, fg=self.accent, cursor="hand2")
            link.pack(anchor="w", pady=2)
            
            # Hover effect
            def on_enter(e):
                link.config(fg=self.text_light)
            def on_leave(e):
                link.config(fg=self.accent)
            def on_click(e):
                self.open_link(url)
            
            link.bind("<Enter>", on_enter)
            link.bind("<Leave>", on_leave)
            link.bind("<Button-1>", on_click)
        
        # Add links
        create_text_link(links_frame, "Reported Issues", 
                        "https://servicemanagementsystems.sharepoint.com/:x:/s/SMARTApplication996-ReleaseReportingTrackingBugs/IQAfDiRhNsshSJmJqwuHThB9AfhMnQYsI1QZy306M0EhHy0?e=rf5TX5")
        
        create_text_link(links_frame, "Book of Work", 
                        "https://servicemanagementsystems-my.sharepoint.com/:x:/r/personal/tdurbin_blockbyblock_com/_layouts/15/Doc.aspx?sourcedoc=%7B2B51D8E9-D3C4-4DEB-A762-F612EE79E359%7D&file=Official%20BOW%20SMART%202026.xlsx&fromShare=true&action=default&mobileredirect=true")
        
        # Theme toggle button (top right)
        toggle_text = "🌙 Dark" if not self.is_dark_mode else "☀️ Light"
        self.theme_toggle_btn = tk.Button(header_frame,
                                          text=toggle_text,
                                          command=self.toggle_theme,
                                          bg=self.bg_card,
                                          fg=self.text_light,
                                          font=("Segoe UI", 10, "bold"),
                                          relief="flat",
                                          cursor="hand2",
                                          padx=20,
                                          pady=10,
                                          borderwidth=2,
                                          highlightthickness=2,
                                          highlightbackground=self.accent,
                                          highlightcolor=self.accent)
        self.theme_toggle_btn.place(relx=1.0, x=-30, y=25, anchor="ne")
        
        # Title
        self.title_label = tk.Label(header_frame, 
                 text="SMART Assistant", 
                 font=("Segoe UI", 36, "bold"),
                 bg=self.bg_dark,
                 fg=self.accent)
        self.title_label.pack(pady=(40, 0))
        
        tk.Label(header_frame, 
                 text="System Management & Reporting Tool  •  v2.0.5", 
                 font=("Segoe UI", 11),
                 bg=self.bg_dark,
                 fg=self.text_dim).pack(pady=(5, 0))
        
        # Main content
        content_frame = tk.Frame(self.frame, bg=self.bg_dark)
        content_frame.pack(fill="both", expand=True, padx=40, pady=40)
        
        # Subtitle
        subtitle_color = "#28a745" if not self.is_dark_mode else self.text_light
        tk.Label(content_frame, 
                 text="Select a tool to get started:", 
                 font=("Segoe UI", 13),
                 bg=self.bg_dark,
                 fg=subtitle_color).pack(pady=(0, 30))
        
        # Button container
        button_frame = tk.Frame(content_frame, bg=self.bg_dark)
        button_frame.pack(expand=True)
        
        # Modern rounded button styling
        button_config = {
            "font": ("Segoe UI", 14, "bold"),
            "relief": "flat",
            "cursor": "hand2",
            "width": 21,  # 25% smaller (was 28)
            "height": 1,  # 25% smaller (was 2, can't be 1.5 so use 1)
            "bd": 0,
            "highlightthickness": 3,
            "highlightbackground": self.bg_dark,
            "activeforeground": "white"
        }
        
        # User Deactivation button - Darker in dark mode
        deactivation_color = "#0066ff" if not self.is_dark_mode else "#0052CC"  # 20% darker
        tk.Button(button_frame, 
                  text="👤 User Deactivation", 
                  command=self.app.show_user_deactivation,
                  bg=deactivation_color,
                  fg="black",
                  activebackground="#0052CC" if not self.is_dark_mode else "#0042A3",
                  **button_config).pack(pady=12, padx=20)
        
        # Add Users submenu button - Green
        add_users_color = "#4CAF50" if not self.is_dark_mode else "#3D8B40"  # 20% darker
        tk.Button(button_frame, 
                  text="➕ Add Users", 
                  command=self.app.show_add_users_menu,
                  bg=add_users_color,
                  fg="black",
                  activebackground="#45a049" if not self.is_dark_mode else "#2F6E33",
                  **button_config).pack(pady=12, padx=20)
        
        # Standards Scans submenu button - Yellow
        standards_color = "#FFD951" if not self.is_dark_mode else "#CCA800"  # 20% darker
        tk.Button(button_frame, 
                  text="📋 Standards Scans", 
                  command=self.app.show_standards_scans_menu,
                  bg=standards_color,
                  fg="black",
                  activebackground="#FFC936" if not self.is_dark_mode else "#B38600",
                  **button_config).pack(pady=12, padx=20)
        
        # Exit button
        exit_frame = tk.Frame(button_frame, bg=self.bg_dark)
        exit_frame.pack(pady=(40, 0), fill="x")
        
        # Exit button - Darker red in dark mode
        exit_color = "#ff4444" if not self.is_dark_mode else "#CC3636"  # 20% darker
        tk.Button(exit_frame, 
                  text="✕ Exit", 
                  command=self.app.on_closing,
                  bg=exit_color,
                  fg="black" if self.is_dark_mode else "white",
                  activebackground="#cc0000" if not self.is_dark_mode else "#A30000",
                  activeforeground="black" if self.is_dark_mode else "white",
                  font=("Segoe UI", 12, "bold"),
                  relief="flat",
                  cursor="hand2",
                  width=18,
                  height=1,
                  bd=0,
                  highlightthickness=2,
                  highlightbackground=self.bg_dark).pack(anchor="center")
        
        # Settings panel - bottom right corner
        settings_frame = tk.Frame(self.frame, bg=self.bg_dark)
        settings_frame.place(relx=1.0, x=-30, rely=1.0, y=-30, anchor="se")
        
        # Settings header
        tk.Label(settings_frame,
                text="⚙️ Settings",
                font=("Segoe UI", 12, "bold"),
                bg=self.bg_dark,
                fg=self.text_light).pack(anchor="w", pady=(0, 15))
        
        # Configure ttk style for themed dropdowns
        from tkinter import ttk
        style = ttk.Style()
        style.theme_use('clam')
        
        # Headless mode dropdown
        headless_frame = tk.Frame(settings_frame, bg=self.bg_dark)
        headless_frame.pack(anchor="w", pady=(0, 5))
        
        tk.Label(headless_frame,
                text="🔧 Chrome Mode:",
                font=("Segoe UI", 10),
                bg=self.bg_dark,
                fg=self.text_light).pack(side="left", padx=(0, 10))
        
        # Style combobox to match theme
        style.configure('Settings.TCombobox',
                       fieldbackground=self.bg_card,
                       background=self.bg_card,
                       foreground=self.text_light,
                       arrowcolor=self.text_light,
                       borderwidth=1)
        style.map('Settings.TCombobox',
                 fieldbackground=[('readonly', self.bg_card)],
                 selectbackground=[('readonly', self.accent)],
                 selectforeground=[('readonly', 'white')])
        
        self.headless_var = tk.StringVar(value="Headless" if self.app.headless_mode else "Visible")
        headless_dropdown = ttk.Combobox(headless_frame,
                                        textvariable=self.headless_var,
                                        values=["Headless", "Visible"],
                                        state="readonly",
                                        width=10,
                                        style='Settings.TCombobox',
                                        font=("Segoe UI", 10))
        headless_dropdown.pack(side="left")
        headless_dropdown.bind("<<ComboboxSelected>>", self.on_headless_change)
        
        # Environment dropdown
        env_frame = tk.Frame(settings_frame, bg=self.bg_dark)
        env_frame.pack(anchor="w", pady=(15, 5))
        
        tk.Label(env_frame,
                text="🌐 Environment:",
                font=("Segoe UI", 10),
                bg=self.bg_dark,
                fg=self.text_light).pack(side="left", padx=(0, 10))
        
        self.env_var = tk.StringVar(value=self.app.environment)
        env_dropdown = ttk.Combobox(env_frame,
                                   textvariable=self.env_var,
                                   values=["Production", "PreProd"],
                                   state="readonly",
                                   width=12,
                                   style='Settings.TCombobox',
                                   font=("Segoe UI", 10))
        env_dropdown.pack(side="left")
        env_dropdown.bind("<<ComboboxSelected>>", self.on_env_change)
        
        # Automation speed dropdown
        speed_label_frame = tk.Frame(settings_frame, bg=self.bg_dark)
        speed_label_frame.pack(anchor="w", pady=(15, 5))
        
        tk.Label(speed_label_frame,
                text="⚡ Automation Speed:",
                font=("Segoe UI", 10),
                bg=self.bg_dark,
                fg=self.text_light).pack(side="left", padx=(0, 10))
        
        # Speed dropdown
        # Speed options - 1x is 1.0 (no modification)
        speed_options = ["0.25x", "0.5x", "0.75x", "1x", "1.25x", "1.5x", "1.75x", "2x"]
        
        # Map display value to actual multiplier
        # Multiplier directly modifies wait times: time.sleep(base_time / multiplier)
        # Higher multiplier = faster (divides by larger number = shorter wait)
        speed_map = {
            "0.25x": 0.25,   # 4x slower
            "0.5x": 0.5,     # 2x slower
            "0.75x": 0.75,   # 1.33x slower
            "1x": 1.0,       # Normal speed (no change)
            "1.25x": 1.25,   # 1.25x faster
            "1.5x": 1.5,     # 1.5x faster
            "1.75x": 1.75,   # 1.75x faster
            "2x": 2.0        # 2x faster
        }
        
        # Find current display speed
        current_display = "1x"
        for display, actual in speed_map.items():
            if abs(self.app.automation_speed - actual) < 0.01:
                current_display = display
                break
        
        self.speed_var = tk.StringVar(value=current_display)
        self.speed_map = speed_map  # Store for callback
        
        speed_dropdown = ttk.Combobox(speed_label_frame,
                                     textvariable=self.speed_var,
                                     values=speed_options,
                                     state="readonly",
                                     width=8,
                                     style='Settings.TCombobox',
                                     font=("Segoe UI", 10))
        speed_dropdown.pack(side="left")
        speed_dropdown.bind("<<ComboboxSelected>>", self.on_speed_change)
        
        # Version label - bottom left corner
        version_label = tk.Label(self.frame,
                                text=f"v{APP_VERSION}",
                                font=("Segoe UI", 9),
                                bg=self.bg_dark,
                                fg=self.text_dim)
        version_label.place(x=30, rely=1.0, y=-20, anchor="sw")
    
    def open_link(self, url):
        """Open URL in browser"""
        import webbrowser
        webbrowser.open(url)
    
    def on_headless_change(self, event=None):
        """Handle headless mode change"""
        self.app.headless_mode = (self.headless_var.get() == "Headless")
        status_text = "✓ Headless ON" if self.app.headless_mode else "✓ Visible Mode"
        print(f"Chrome mode changed: {status_text}")
        self.app.save_settings()
    
    def on_env_change(self, event=None):
        """Handle environment change"""
        self.app.environment = self.env_var.get()
        self.app.update_login_url()
        logger.info(f"Environment changed to: {self.app.environment}")
        print(f"Environment changed to: {self.app.environment}")
        self.app.save_settings()
    
    def toggle_headless_mode(self):
        """Toggle Chrome headless mode (deprecated - using dropdown now)"""
        self.app.headless_mode = self.headless_var.get()
        status_text = "✓ Headless ON" if self.app.headless_mode else "✓ Headless OFF"
        print(f"Chrome mode changed: {status_text}")
        self.app.save_settings()
    
    def on_speed_change(self, event=None):
        """Handle automation speed change"""
        speed_str = self.speed_var.get()
        # Use the speed map to get actual multiplier
        self.app.automation_speed = self.speed_map.get(speed_str, 1.2)
        logger.info(f"Automation speed changed to: {speed_str} (actual: {self.app.automation_speed})")
        print(f"Automation speed changed to: {speed_str} (actual: {self.app.automation_speed})")
        # Save settings
        self.app.save_settings()
    
    def toggle_theme(self):
        """Toggle theme and reload menu"""
        self.app.is_dark_mode = not self.app.is_dark_mode
        self.app.apply_theme_colors()
        
        # Update title bar
        self.app.enable_dark_title_bar(self.app.root)
        
        # Update root background
        self.app.root.configure(bg=self.app.bg_dark)
        
        # Save all settings (theme + headless mode)
        self.app.save_settings()
        
        # Reload menu
        self.app.show_main_menu()
    
    def destroy(self):
        self.frame.destroy()


# ======================================================================
# MAIN FUNCTION
# ======================================================================

def main():
    if sys.platform == "win32":
        import ctypes
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    
    app = SMARTAssistantApp()
    app.run()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        logging.error("Fatal error:")
        logging.error(traceback.format_exc())
        save_crash_report(str(e))
        messagebox.showerror("Fatal Error", f"SMART Assistant error:\n\n{str(e)}")



# ======================================================================
# USER DEACTIVATION CORE FUNCTIONALITY
# ======================================================================

